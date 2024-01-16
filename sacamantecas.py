#! /usr/bin/env python
"""See "README.md" for details."""
import sys
if sys.platform != 'win32':
    print('\nThis application is compatible only with the Win32 platform.')
    sys.exit(None)

# pylint: disable=wrong-import-position
import atexit
import configparser
from ctypes import byref, c_uint, create_unicode_buffer, windll
from ctypes.wintypes import MAX_PATH as MAX_PATH_LEN
from enum import auto, IntEnum, StrEnum
import errno
from http.client import HTTPException
from html.parser import HTMLParser
import logging
from logging.config import dictConfig
from msvcrt import get_osfhandle, getch
from pathlib import Path
import platform
import re
from shutil import copy2
from textwrap import dedent
import time
import traceback as tb
from types import SimpleNamespace
from urllib.error import HTTPError, URLError
from urllib.parse import quote, unquote, urlparse, urlunparse
from urllib.request import urlopen, Request
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_STRING as CELLTYPE_STRING
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter

from version import DEVELOPMENT_MODE, SEMVER


APP_NAME = Path(__file__).stem

class Messages(StrEnum):
    """Messages."""
    INITIALIZATION_ERROR = 'Error de inicialización de la aplicación.'

    PRESS_ANY_KEY = '\nPulse cualquier tecla para continuar...'
    KEYBOARD_INTERRUPT = 'El usuario interrumpió la operación de la aplicación.'
    NO_ARGUMENTS = dedent('''
        No se han especificado fuentes de entrada para ser procesadas.

        Arrastre y suelte un fichero de entrada sobre el icono de la aplicación,
        o bien proporcione los nombres de las fuentes de entrada como argumentos.
    ''').lstrip()

    DEBUGGING_INIT = 'Registro de depuración iniciado.'
    APP_BANNER = f'{APP_NAME} versión {SEMVER}'
    PROCESS_DONE = '\nProceso finalizado.'
    DEBUGGING_DONE = 'Registro de depuración finalizado.'

    ERROR_HEADER = f'\n*** Error en {APP_NAME}.\n'
    WARNING_HEADER = '* Aviso: '
    ERROR_DETAILS_HEADING = '\nInformación adicional sobre el error:'
    ERROR_DETAILS_TAIL_CHAR = '·'

    UNEXPECTED_OSERROR = 'Error inesperado del sistema operativo.'
    OSERROR_DETAILS = 'type = {}\nerrno = {}\nwinerror = {}\nstrerror = {}\nfilename = {}\nfilename2 = {}'
    UNHANDLED_EXCEPTION = 'Excepción sin gestionar.'
    EXCEPTION_DETAILS = 'type = {}\nvalue = {}\nargs: {}'
    EXCEPTION_DETAILS_ARG = '\n  [{}] {}'
    TRACEBACK_HEADER = '\n\ntraceback:\n{}'
    TRACEBACK_FRAME_HEADER = '▸ {}\n'
    TRACEBACK_FRAME_LINE = '  {}, {}: {}\n'
    TRACEBACK_TOPLEVEL_FRAME = '<module>'
    UNKNOWN_ERRNO = 'desconocido'

    EMPTY_PROFILES = 'No hay perfiles definidos en el fichero de perfiles «{}».'
    MISSING_PROFILES = 'No se encontró o no se pudo leer el fichero de perfiles «{}».'
    PROFILES_WRONG_SYNTAX = 'Error de sintaxis «{}» leyendo el fichero de perfiles.'
    PROFILES_WRONG_SYNTAX_DETAILS = 'Perfil «{}», {}:\n  {}{}{}\n  {:_<{}}^'
    PROFILES_WRONG_SYNTAX_DETAILS_SEPARATOR = ' = '
    INVALID_PROFILE = 'El perfil «{}» es inválido.'
    PROFILE_WITHOUT_URL = 'El perfil no incluye un patrón de URL.'
    UNKNOWN_URL_TYPE = 'El URL «{}» es de tipo desconocido.'
    NO_MATCHING_PROFILE = 'No se encontró un perfil para procesar el URL.'

    OSLIKE_URLERROR = 'Error de red {}: {}.'
    HTTP_PROTOCOL_URLERROR = 'Error de protocolo HTTP {}: {}.'
    GENERIC_URLERROR = 'Error de URL{}: {}.'
    URL_ACCESS_ERROR = 'No resultó posible acceder a la dirección especificada.'
    HTTP_RETRIEVAL_ERROR = 'No se obtuvieron contenidos.'
    CONNECTION_ERROR = 'Se produjo un error de conexión «{}» accediendo al URL.'
    NO_CONTENTS_ERROR = 'No se recibieron contenidos del URL.'

    SKIMMING_MARKER = '\nSacando las mantecas:\n'
    SOURCE_LABEL = 'Fuente: {}'
    UNSUPPORTED_SOURCE = 'La fuente no es de un tipo admitido.'
    INPUT_FILE_INVALID = 'El fichero de entrada es inválido ({}).'
    INPUT_FILE_NOT_FOUND = 'No se encontró el fichero de entrada.'
    INPUT_FILE_NO_PERMISSION = 'No hay permisos suficientes para leer el fichero de entrada.'
    OUTPUT_FILE_NO_PERMISSION = 'No hay permisos suficientes para crear el fichero de salida.'
    NO_METADATA_FOUND = 'No se obtuvieron metadatos.'


class Debug(StrEnum):
    """Debugging messages."""
    PROCESSING_ARG = 'Procesando argumento «{}».'
    ARG_IS_SOURCE_SINGLE_URL = 'El argumento es una fuente de tipo single_url.'
    ARG_IS_SOURCE_TEXTFILE = 'El argumento es una fuente de tipo textfile.'
    ARG_IS_SOURCE_SPREADSHEET = 'El argumento es una fuente de tipo spreadsheet.'
    ARG_IS_SOURCE_UNSUPPORTED = 'El argumento no es un tipo de fuente admitido.'

    LOADING_PROFILES = 'Obteniendo perfiles desde «{}».'
    FOUND_PROFILES = 'Se obtuvieron los siguientes perfiles: {}.'
    DETECTED_PROFILE = 'Perfil detectado: «{}».'

    PROCESSING_URL = 'Procesando URL «{}».'
    REDIRECTED_URL = 'URL redirigido a «{}».'

    CHARSET_NOT_IN_HEADERS = 'Charset no detectado en las cabeceras.'
    CHARSET_IN_HEADERS = 'Charset detectado en las cabeceras.'

    CHARSET_FROM_HTTP_EQUIV = 'Charset detectado mediante meta http-equiv.'
    CHARSET_FROM_META_CHARSET = 'Charset detectado mediante meta charset.'
    CHARSET_FROM_DEFAULT = 'Charset not detectado, usando valor por defecto.'

    CONTENTS_ENCODING = 'Contenidos codificados con charset «{}».'

    HTML_START_TAG = '➜ HTML <{}{}{}>'
    METADATA_KEY_FOUND = 'Se encontró la clave «{}».'
    METADATA_VALUE_FOUND = 'Se encontró el valor «{}».'
    METADATA_IS_EMPTY = 'Metadato vacío.'
    METADATA_MISSING_VALUE = 'Metadato «{}» incompleto, ignorando.'
    METADATA_MISSING_KEY = 'No se encontró una clave, usando «{}».'
    METADATA_OK = 'Metadato correcto «{}: {}».'
    METADATA_KEY_MARKER_FOUND = 'Se encontró una marca de clave «{}».'
    METADATA_VALUE_MARKER_FOUND = 'Se encontró una marca de valor «{}».'
    METADATA_MARKER_FOUND = 'Se encontró una marca de metadato «{}».'

    COPYING_WORKBOOK = 'Copiando workbook a «{}».'
    WORKING_SHEET = 'La hoja con la que se trabajará es «{}»".'
    INSERTING_HEADING_ROW = 'Insertando fila de cabeceras.'
    PROCESSING_ROW = 'Procesando fila {}.'
    NONSTRING_CELL = 'La celda «{}» no es de tipo cadena, será ignorada.'
    URL_FOUND_IN_CELL = 'Se encontró un URL en la celda «{}»: {}'
    NEW_METADATA_FOUND = 'Se encontró un metadato nuevo, «{}».'
    METADATA_STORED_IN_COLUMN = 'El metadato «{}» irá en la columna «{}».'

    DUMPING_METADATA_TO_SINK = 'Volcando metadatos a «{}».'
    DUMPING_METADATA_K_V = 'Añadiendo metadato «{}» con valor «{}».'

    PARSER_NESTING_ERROR_K_IN_V = 'Problema de anidación (clave dentro de valor), restableciendo parser.'
    PARSER_NESTING_ERROR_V_IN_K = 'Problema de anidación (valor dentro de clave), restableciendo parser.'


class ExitCodes(IntEnum):
    """Standardized exit codes."""
    SUCCESS = 0
    NO_ARGUMENTS = 1
    WARNING = 2
    ERROR = 3
    KEYBOARD_INTERRUPT = 127


try:
    if getattr(sys, 'frozen', False):
        ROOT_PATH = sys.executable
    else:
        ROOT_PATH = __file__
except NameError:
    print(Messages.INITIALIZATION_ERROR, file=sys.stderr)
    sys.exit(ExitCodes.ERROR)
ROOT_PATH = Path(ROOT_PATH).resolve().parent


# Some constants used to prevent mistyping.
EMPTY_STRING = ''
UTF8 = 'utf-8'
ASCII = 'ascii'


class Config():  # pylint: disable=too-few-public-methods
    """Application configuration values."""
    TIMESTAMP_FORMAT = '%Y%m%d_%H%M%S'
    USER_AGENT = ' '.join(dedent(f'''
        {APP_NAME}/{SEMVER}
        +https://github.com/DervishD/sacamantecas
        (Windows {platform.version()};
        {platform.architecture()[0]};
        {platform.machine()})
    ''').splitlines()).lstrip()

    ACCEPTED_URL_SCHEMES = ('https', 'http', 'file')

    FALLBACK_HTML_CHARSET = 'ISO-8859-1'

    PROFILE_URL_PATTERN_KEY = 'url'

    TEXTFILE_SUFFIX = '.txt'
    SPREADSHEET_SUFFIX = '.xlsx'

    TIMESTAMP_STEM = time.strftime(f'_{TIMESTAMP_FORMAT}')
    SINKFILE_STEM = '_out'

    LOGFILE_PATH = ROOT_PATH / f'{APP_NAME}_log{"" if DEVELOPMENT_MODE else TIMESTAMP_STEM}{TEXTFILE_SUFFIX}'
    DEBUGFILE_PATH = ROOT_PATH / f'{APP_NAME}_debug{"" if DEVELOPMENT_MODE else TIMESTAMP_STEM}{TEXTFILE_SUFFIX}'
    INIFILE_PATH = ROOT_PATH / f'{APP_NAME}.ini'

    LOGGING_INDENTCHAR = ' '
    LOGGING_FORMAT_STYLE = '{'
    LOGGING_FALLBACK_FORMAT = '{message}'
    LOGGING_DEBUGFILE_FORMAT = '{asctime}.{msecs:04.0f} {levelname}| {message}'
    LOGGING_LOGFILE_FORMAT = '{asctime} {message}'
    LOGGING_CONSOLE_FORMAT = '{message}'

    TEXTSINK_METADATA_HEADER = '{}\n'
    TEXTSINK_METADATA_INDENT = '  '
    TEXTSINK_METADATA_SEPARATOR = ': '
    TEXTSINK_METADATA_PAIR = f'{TEXTSINK_METADATA_INDENT}{{}}{TEXTSINK_METADATA_SEPARATOR}{{}}\n'
    TEXTSINK_METADATA_FOOTER = '\n'

    SPREADSHEET_METADATA_COLUMN_MARKER = '[sm] '
    SPREADSHEET_METADATA_COLUMN_TITLE = f'{SPREADSHEET_METADATA_COLUMN_MARKER}{{}}'
    SPREADSHEET_CELL_FONT = 'Calibri'
    SPREADSHEET_CELL_COLOR = 'baddad'
    SPREADSHEET_CELL_FILL = 'solid'

    URL_UNSAFE_CHARS_RE = r'\W'
    URL_UNSAFE_REPLACE_CHAR = '_'

    FILE_URL_SAFE_CHARS = ':/'

    META_HTTP_EQUIV_CHARSET_RE = rb'<meta http-equiv="content-type".*charset="([^"]+)"'
    META_CHARSET_RE = rb'<meta charset="([^"]+)"'
    META_REFRESH_RE = rb'<meta http-equiv="refresh" content="(?:[^;]+;\s+)?URL=([^"]+)"'


# Needed for having VERY basic logging when the code is imported rather than run.
logging.basicConfig(
    level=logging.NOTSET,
    style=Config.LOGGING_FORMAT_STYLE,
    format=Config.LOGGING_FALLBACK_FORMAT,
    force=True
)
logging.indent = lambda level=None: None
logging.dedent = lambda level=None: None


# Reconfigure standard output streams so they use UTF-8 encoding, even if
# they are redirected to a file when running the application from a shell.
sys.stdout.reconfigure(encoding=UTF8)
sys.stderr.reconfigure(encoding=UTF8)


class BaseApplicationError(Exception):
    """Base class for all custom application exceptions."""
    # cSpell:ignore vararg
    def __init__ (self, message, details=EMPTY_STRING, *args, **kwargs):  # pylint: disable=keyword-arg-before-vararg
        self.details = details
        super().__init__(message, *args, **kwargs)

class ProfilesError(BaseApplicationError):
    """Raise for profile-related errors."""

class SourceError(BaseApplicationError):
    """Raise for source-related errors."""

class SkimmingError(BaseApplicationError):
    """Raise for skimming-related errors."""


class Profile():  # pylint: disable=too-few-public-methods
    """Abstraction for profiles."""
    def __init__(self, url_pattern, parser, parser_config):
        self.url_pattern = url_pattern
        self.parser = parser
        self.parser_config = parser_config
    def __eq__(self, other):
        if not isinstance(other, self.__class__):
            return NotImplemented
        if self.url_pattern != other.url_pattern or self.parser_config != other.parser_config:
            return False
        return type(self.parser).__name__ == type(other.parser).__name__


class BaseParser(HTMLParser):
    """Base class for catalogue parsers."""
    PARAMETERS = set()
    EMPTY_KEY_PLACEHOLDER = '[vacío]'
    MULTIDATA_SEPARATOR = ' / '
    MULTIVALUE_SEPARATOR = ' === '

    def __init__(self, *args, **kwargs):
        """Initialize object."""
        self.within_k = None
        self.within_v = None
        self.current_k = None
        self.current_v = None
        self.last_k = None
        self.retrieved_metadata = None
        self.config = None
        super().__init__(*args, **kwargs)

    def reset(self):
        """Reset parser state. Called implicitly from __init__()."""
        super().reset()
        self.within_k = self.within_v = False
        self.current_k = self.current_v = self.last_k = EMPTY_STRING
        self.retrieved_metadata = {}

    def handle_starttag(self, tag, attrs):
        """Handle opening tags."""
        logging.debug(Debug.HTML_START_TAG.format(tag, ' ' * bool(attrs), ' '.join((f'{k}="{v}"' for k, v in attrs))))

    def handle_data(self, data):
        """Handle data."""
        if self.within_k or self.within_v:
            # Clean up the received data by removing superfluous whitespace
            # characters, including newlines, carriage returns, etc.
            data = ' '.join(data.split())
            if not data:
                return
        if self.within_k:
            logging.debug(Debug.METADATA_KEY_FOUND.format(data))
            self.current_k += data.rstrip(':')
            self.last_k = self.current_k
            return
        if self.within_v:
            logging.debug(Debug.METADATA_VALUE_FOUND.format(data))
            self.current_v = f'{self.current_v}{self.MULTIDATA_SEPARATOR if self.current_v else EMPTY_STRING}{data}'
            return

    def configure(self, config):
        """
        Set up the parser with a different configuration, that is, a different
        set of values for the suppported parameters.

        Only supported config parameters are used, the rest are ignored.

        This operation also resets the parser.
        """
        self.reset()
        self.config = {key: value for key, value in config.items() if key in self.PARAMETERS}

    def store_metadata(self):
        """Store found metadata, handling missing parts."""
        if not self.current_k and not self.current_v:
            logging.debug(Debug.METADATA_IS_EMPTY)
        if self.current_k and not self.current_v:
            logging.debug(Debug.METADATA_MISSING_VALUE.format(self.current_k))
        if not self.current_k and self.current_v:
            self.current_k = self.last_k if self.last_k else self.EMPTY_KEY_PLACEHOLDER
            logging.debug(Debug.METADATA_MISSING_KEY.format(self.current_k))
        if self.current_k and self.current_v:
            if self.current_k not in self.retrieved_metadata:
                self.retrieved_metadata[self.current_k] = []
            # A set is not used instead of the code below, to preserve order.
            if self.current_v not in self.retrieved_metadata[self.current_k]:
                self.retrieved_metadata[self.current_k].append(self.current_v)
            logging.debug(Debug.METADATA_OK.format(self.current_k, self.current_v))
        self.current_k = self.current_v = EMPTY_STRING

    def get_metadata(self):
        """Get retrieved metadata so far."""
        metadata = {}
        for key, value in self.retrieved_metadata.items():
            metadata[key] = self.MULTIVALUE_SEPARATOR.join(value)
        return metadata


class OldRegimeParser(BaseParser):  # pylint: disable=unused-variable
    """
    Parser for Old Regime catalogues which use different HTML class attributes
    to mark metadata keys and metadata values.

    This is inherently complex, specially when dealing with ill-formed HTML.

    So, in order to keep this parser as simple as possible, some assumptions are
    made. See the comments below to know which those are.
    """
    K_CLASS = 'k_class'
    V_CLASS = 'v_class'
    PARAMETERS = BaseParser.PARAMETERS | {K_CLASS, V_CLASS}
    CLASS_ATTR = 'class'

    def __init__(self, *args, **kwargs):
        """Initialize object."""
        self.current_k_tag = None
        self.current_v_tag = None
        super().__init__(*args, **kwargs)

    def reset(self):
        """Reset parser state. Called implicitly from __init__()."""
        super().reset()
        self.current_k_tag = self.current_v_tag = EMPTY_STRING

    def handle_starttag(self, tag, attrs):
        """Handle opening tags."""
        super().handle_starttag(tag, attrs)
        for attr in attrs:
            if attr[0] == self.CLASS_ATTR and (match := self.config[self.K_CLASS].search(attr[1])):
                logging.debug(Debug.METADATA_KEY_MARKER_FOUND.format(match.group(0)))
                self.within_k = True
                self.current_k = EMPTY_STRING
                self.current_k_tag = tag
                if self.within_v:
                    # If still processing a value, notify about the nesting error
                    # but reset parser so everything starts afresh, like if a new
                    # key had been found.
                    logging.debug(Debug.PARSER_NESTING_ERROR_K_IN_V)
                    self.within_v = False
                    self.current_v = EMPTY_STRING
                    self.current_v_tag = None
                break
            if attr[0] == self.CLASS_ATTR and (match := self.config[self.V_CLASS].search(attr[1])):
                logging.debug(Debug.METADATA_VALUE_MARKER_FOUND.format(match.group(0)))
                self.within_v = True
                self.current_v = EMPTY_STRING
                self.current_v_tag = tag
                if self.within_k:
                    # If still processing a key, the nesting error can still be
                    # recovered up to a certain point. If some data was got for
                    # the key, the parser is left in within_v mode to try to get
                    # the corresponding value. Otherwise the parser is reset.
                    logging.debug(Debug.PARSER_NESTING_ERROR_V_IN_K)
                    self.within_k = False
                    self.current_k_tag = None
                    if not self.current_k:
                        self.within_v = False
                        self.current_v_tag = None
                break

    def handle_endtag(self, tag):
        """Handle closing tags."""
        super().handle_endtag(tag)
        if self.within_k and tag == self.current_k_tag:
            self.within_k = False
            self.current_k_tag = None
            return
        if self.within_v and tag == self.current_v_tag:
            self.within_v = False
            self.current_v_tag = None
            self.store_metadata()
            return


# cSpell:ignore Baratz
class BaratzParser(BaseParser):   # pylint: disable=unused-variable
    """
    Parser for catalogues whose contents have been generated by the new Baratz
    frontend, which does not use the class attributes to mark metadata keys and
    values. Instead they have an HTML element with acts as a start of metadata
    marker, containing the metadata as a list of <dt>/<dd> pairs.

    Within that list, the <dt> HTML element contains the metadata key, whereas
    the <dd> HTML element containing the metadata value.

    The HTML element which signals the start of a metadata block is defined by
    its HTML tag, the attribute containing the marker value and of course the
    marker value itself.
    """
    M_TAG = 'm_tag'
    M_ATTR = 'm_attr'
    M_VALUE = 'm_value'
    PARAMETERS = BaseParser.PARAMETERS | {M_TAG, M_ATTR, M_VALUE}
    K_TAG = 'dt'
    V_TAG = 'dd'

    def __init__(self, *args, **kwargs):
        """Initialize object."""
        self.within_meta = None
        super().__init__(*args, **kwargs)

    def reset(self):
        """Reset parser state. Called implicitly from __init__()."""
        super().reset()
        self.within_meta = False

    def handle_starttag(self, tag, attrs):
        """Handle opening tags."""
        super().handle_starttag(tag, attrs)
        if not self.within_meta:
            if not self.config[self.M_TAG].fullmatch(tag):
                return
            for attr in attrs:
                if self.config[self.M_ATTR].fullmatch(attr[0]) and self.config[self.M_VALUE].search(attr[1]):
                    logging.debug(Debug.METADATA_MARKER_FOUND.format(attr[1]))
                    self.within_meta = True
                    return
        else:
            if tag == self.K_TAG:
                logging.debug(Debug.METADATA_KEY_MARKER_FOUND.format(tag))
                self.within_k = True
                if self.within_v:
                    # If still processing a value, notify about the nesting error
                    # but reset parser so everything starts afresh, like if a new
                    # key had been found.
                    logging.debug(Debug.PARSER_NESTING_ERROR_K_IN_V)
                    self.within_v = False
                    self.current_v = EMPTY_STRING
                return
            if tag == self.V_TAG:
                logging.debug(Debug.METADATA_VALUE_MARKER_FOUND.format(tag))
                self.within_v = True
                if self.within_k:
                    # If still processing a key, the nesting error can still be
                    # recovered up to a certain point. If some data was got for
                    # the key, the parser is left in within_v mode to try to get
                    # the corresponding value. Otherwise the parser is reset.
                    logging.debug(Debug.PARSER_NESTING_ERROR_V_IN_K)
                    self.within_k = False
                    if not self.current_k:
                        self.within_v = False
                return

    def handle_endtag(self, tag):
        """Handle closing tags."""
        super().handle_endtag(tag)
        if self.within_meta and self.config[self.M_TAG].fullmatch(tag):
            self.within_meta = False
            return
        if self.within_k and tag == self.K_TAG:
            self.within_k = False
            return
        if self.within_v and tag == self.V_TAG:
            self.within_v = False
            self.store_metadata()
            return


def error(message, details=EMPTY_STRING):
    """Helper for preprocessing error messages."""
    message = str(message)
    details = str(details)
    logging.indent(0)
    logging.error(Messages.ERROR_HEADER)
    logging.indent(len(Messages.ERROR_HEADER.lstrip().split(' ', maxsplit=1)[0]) + 1)
    logging.error(message)
    if details.strip():
        logging.error(Messages.ERROR_DETAILS_HEADING)
        logging.error('\n'.join(f'| {line}' for line in details.splitlines()))
        logging.error(Messages.ERROR_DETAILS_TAIL_CHAR)
    logging.indent(0)


def warning(message):
    """Helper for prepending a header to warning messages."""
    message = str(message)
    message = Messages.WARNING_HEADER + message[0].lower() + message[1:]
    logging.warning(message)


def is_accepted_url(value):
    """Check if value is an accepted URL or not."""
    # The check is quite crude but works for the application's needs.
    try:
        return urlparse(value).scheme in Config.ACCEPTED_URL_SCHEMES
    except ValueError:
        return False


def generate_sink_filename(base_filename):
    """
    Generate a filename usable as data sink, based upon base_filename.
    """
    return base_filename.with_stem(base_filename.stem + Config.SINKFILE_STEM)


class WFKStatuses(IntEnum):
    """Return statuses for wait_for_keypress()."""
    IMPORTED = auto()
    NO_CONSOLE_ATTACHED = auto()
    NO_CONSOLE_TITLE = auto()
    NO_TRANSIENT_FROZEN = auto()
    NO_TRANSIENT_PYTHON = auto()
    WAIT_FOR_KEYPRESS = auto()

def wait_for_keypress():
    """Wait for a keypress to continue if sys.stdout is a real console AND the console is transient."""
    # First of all, if this script is being imported rather than run,
    # then the application must NOT pause. Absolutely NOT.
    if __name__ != '__main__':
        return WFKStatuses.IMPORTED

    # If no console is attached, then the application must NOT pause.
    #
    # Since sys.stdout.isatty() returns True under Windows when sys.stdout
    # is redirected to NUL, another (more complex) method, is needed here.
    # The test below has been adapted from https://stackoverflow.com/a/33168697
    if not windll.kernel32.GetConsoleMode(get_osfhandle(sys.stdout.fileno()), byref(c_uint())):
        return WFKStatuses.NO_CONSOLE_ATTACHED

    # If there is a console attached, the application must pause ONLY if that
    # console will automatically close when the application finishes, hiding
    # any messages printed by the application. In other words, pause only if
    # the console is transient.
    #
    # Determining if a console is transient is not easy as there is no
    # bulletproof method available for every possible circumstance.
    #
    # There are TWO main scenarios: a frozen executable and a .py file.
    # In both cases, the console title has to be obtained.
    buffer_size = MAX_PATH_LEN + 1
    console_title = create_unicode_buffer(buffer_size)
    if not windll.kernel32.GetConsoleTitleW(console_title, buffer_size):
        return WFKStatuses.NO_CONSOLE_TITLE
    console_title = console_title.value

    # If the console is not transient, return, do not pause.
    #
    # For a frozen executable, it is more or less easy: if the console title
    # is not equal to sys.executable, then the console is NOT transient.
    #
    # For a .py file, it is a bit more complicated, but in most cases if the
    # console title contains the name of the .py file, the console is NOT a
    # transient console.
    if getattr(sys, 'frozen', False):
        if console_title != sys.executable:
            return WFKStatuses.NO_TRANSIENT_FROZEN
    elif APP_NAME in console_title:
        return WFKStatuses.NO_TRANSIENT_PYTHON

    print(Messages.PRESS_ANY_KEY, end=EMPTY_STRING, flush=True)
    getch()
    return WFKStatuses.WAIT_FOR_KEYPRESS


def excepthook(exc_type, exc_value, exc_traceback):
    """Handle unhandled exceptions, default exception hook."""
    if isinstance(exc_value, OSError):
        message = Messages.UNEXPECTED_OSERROR
        details = Messages.OSERROR_DETAILS.format(
            exc_type.__name__,
            EMPTY_STRING if exc_value.errno is None else errno.errorcode[exc_value.errno],
            EMPTY_STRING if exc_value.winerror is None else exc_value.winerror,
            exc_value.strerror,
            EMPTY_STRING if exc_value.filename is None else exc_value.filename,
            EMPTY_STRING if exc_value.filename2 is None else exc_value.filename2,
        )
    else:
        message = Messages.UNHANDLED_EXCEPTION
        args = EMPTY_STRING
        for arg in exc_value.args:
            args += Messages.EXCEPTION_DETAILS_ARG.format(type(arg).__name__, arg)
        details = Messages.EXCEPTION_DETAILS.format(exc_type.__name__, str(exc_value), args)
    current_filename = None
    traceback = EMPTY_STRING
    for frame in tb.extract_tb(exc_traceback):
        if current_filename != frame.filename:
            traceback += Messages.TRACEBACK_FRAME_HEADER.format(frame.filename)
            current_filename = frame.filename
        frame.name = APP_NAME if frame.name == Messages.TRACEBACK_TOPLEVEL_FRAME else frame.name
        traceback += Messages.TRACEBACK_FRAME_LINE.format(frame.lineno, frame.name, frame.line)
    details += Messages.TRACEBACK_HEADER.format(traceback) if traceback else EMPTY_STRING
    error(message, details)


def loggerize(function):
    """Decorator which enables logging for function."""
    def loggerize_wrapper(*args, **kwargs):
        setup_logging(Config.LOGFILE_PATH, Config.DEBUGFILE_PATH)

        logging.debug(Messages.DEBUGGING_INIT)
        logging.info(Messages.APP_BANNER)
        logging.debug(Config.USER_AGENT)

        status = function(*args, **kwargs)

        logging.info(Messages.PROCESS_DONE)
        logging.debug(Messages.DEBUGGING_DONE)
        logging.shutdown()
        return status
    return loggerize_wrapper


def setup_logging(log_filename, debug_filename):
    """
    Sets up logging system, disabling all existing loggers.

    With the current configuration ALL logging messages are sent to the debug
    file and messages with levels over logging.INFO are sent to the log file.

    Also, logging.INFO messages are sent to sys.stdout, without a timestamp.
    Finally, messages with levels over logging.INFO are sent to sys.stderr, also
    without a timestamp.
    """
    class CustomFormatter(logging.Formatter):
        """Simple custom formatter for logging messages."""
        def format(self, record):
            """
            Format multiline records so they look like multiple records.
            Indent message according to current indentation level.
            """
            message = super().format(record)
            preamble, message = message.partition(record.message)[:2]
            return '\n'.join([f'{preamble}{record.indent}{line.strip()}'.rstrip() for line in message.splitlines()])

    logging_configuration = {
        'version': 1,
        'disable_existing_loggers': True,
        'formatters': {
            'debugfile_formatter': {
                '()': CustomFormatter,
                'style': Config.LOGGING_FORMAT_STYLE,
                'format': Config.LOGGING_DEBUGFILE_FORMAT,
                'datefmt': Config.TIMESTAMP_FORMAT,
            },
            'logfile_formatter': {
                '()': CustomFormatter,
                'style': Config.LOGGING_FORMAT_STYLE,
                'format': Config.LOGGING_LOGFILE_FORMAT,
                'datefmt': Config.TIMESTAMP_FORMAT,
            },
            'console_formatter': {
                '()': CustomFormatter,
                'style': Config.LOGGING_FORMAT_STYLE,
                'format': Config.LOGGING_CONSOLE_FORMAT,
            },
        },
        'filters': {
            'debugfile_filter': {'()': lambda: lambda record: record.msg.strip() and record.levelno > logging.NOTSET},
            'logfile_filter': {'()': lambda: lambda record: record.msg.strip() and record.levelno >= logging.INFO},
            'stdout_filter': {'()': lambda: lambda record: record.msg.strip() and record.levelno == logging.INFO},
            'stderr_filter': {'()': lambda: lambda record: record.msg.strip() and record.levelno > logging.INFO},
        },
        'handlers': {
            'debugfile_handler': {
                'level': logging.NOTSET,
                'formatter': 'debugfile_formatter',
                'filters': ['debugfile_filter'],
                'class': logging.FileHandler,
                'filename': debug_filename,
                'mode': 'w',
                'encoding': UTF8,
            },
            'logfile_handler': {
                'level': logging.NOTSET,
                'formatter': 'logfile_formatter',
                'filters': ['logfile_filter'],
                'class': logging.FileHandler,
                'filename': log_filename,
                'mode': 'w',
                'encoding': UTF8,
            },
            'stdout_handler': {
                'level': logging.NOTSET,
                'formatter': 'console_formatter',
                'filters': ['stdout_filter'],
                'class': logging.StreamHandler,
                'stream': sys.stdout,
            },
            'stderr_handler': {
                'level': logging.NOTSET,
                'formatter': 'console_formatter',
                'filters': ['stderr_filter'],
                'class': logging.StreamHandler,
                'stream': sys.stderr,
            },
        },
        'loggers': {
            '': {
                'level': logging.NOTSET,
                'handlers': [
                    'debugfile_handler',
                    'logfile_handler',
                    'stdout_handler',
                    'stderr_handler'
                ],
                'propagate': False,
            },
        },
    }

    dictConfig(logging_configuration)

    setattr(logging.getLogger(), 'indentlevel', 0)

    current_factory = logging.getLogRecordFactory()
    levelname_template = f'{{:{len(max(logging.getLevelNamesMapping(), key=len))}}}'
    def record_factory(*args, **kwargs):
        """LogRecord factory which supports indentation."""
        record = current_factory(*args, **kwargs)
        record.indent = Config.LOGGING_INDENTCHAR * logging.getLogger().indentlevel
        record.levelname = levelname_template.format(record.levelname)
        return record
    logging.setLogRecordFactory(record_factory)

    increase_indent_symbol = '+'
    decrease_indent_symbol = '-'
    def set_indent_level(level):
        """
        Set current indentation level.

        If level is increase_indent_symbol, current indentation level is increased.
        If level is decrease_indent_symbol, current indentation level is decreased.
        For any other value, indentation level is set to the provided value.
        """
        if level == '+':
            logging.getLogger().indentlevel += 1
            return
        if level == '-':
            logging.getLogger().indentlevel -= 1
            return
        logging.getLogger().indentlevel = level
    # Both logging.indent() and logging.dedent() support a parameter specifying an
    # exact FINAL indentation level, not an indentation increment/decrement!
    # These two helpers are provided in order to improve readability, since the
    # set_logging_indent_level() function can be used directly.
    logging.indent = lambda level=None: set_indent_level(increase_indent_symbol if level is None else level)
    logging.dedent = lambda level=None: set_indent_level(decrease_indent_symbol if level is None else level)


def keyboard_interrupt_handler(function):
    """Decorator which wraps function with a simple KeyboardInterrupt handler."""
    def handle_keyboard_interrupt_wrapper(*args, **kwargs):
        try:
            return function(*args, **kwargs)
        except KeyboardInterrupt:
            warning(Messages.KEYBOARD_INTERRUPT)
            return ExitCodes.KEYBOARD_INTERRUPT
    return handle_keyboard_interrupt_wrapper


def load_profiles(filename):
    """
    Load the profiles from filename.

    Return the preprocessed list of profiles as a dictionary where the keys are
    the profiles which were found in filename and the values are dictionaries
    containing the corresponding profile configuration items as key-value pairs.

    Raise MissingProfilesError if filename cannot be opened or read.

    Raise ProfilesSyntaxError if there is any syntax error in filename.

    The returned dictionary will be empty if no profiles or only empty profiles
    are present in filename.
    """
    config = configparser.ConfigParser()
    logging.debug(Debug.LOADING_PROFILES.format(filename))
    try:
        with open(filename, encoding=UTF8) as inifile:
            config.read_file(inifile)
    except (FileNotFoundError, PermissionError) as exc:
        raise ProfilesError(Messages.MISSING_PROFILES.format(exc.filename)) from exc
    except configparser.Error as exc:
        errorname = type(exc).__name__.removesuffix('Error')
        raise ProfilesError(Messages.PROFILES_WRONG_SYNTAX.format(errorname), exc) from exc

    profiles = {}
    parsers = [parser() for parser in BaseParser.__subclasses__()]
    for section in config.sections():
        if not config[section]:
            continue
        parser_config = {}
        for key, value in config[section].items():
            if not value:
                continue
            try:
                parser_config[key] = re.compile(value, re.IGNORECASE)
            except re.error as exc:
                details = Messages.PROFILES_WRONG_SYNTAX_DETAILS.format(
                    section, exc.msg,
                    key, Messages.PROFILES_WRONG_SYNTAX_DETAILS_SEPARATOR, exc.pattern,
                    EMPTY_STRING, exc.pos + len(key) + len(Messages.PROFILES_WRONG_SYNTAX_DETAILS_SEPARATOR)
                )
                raise ProfilesError(Messages.PROFILES_WRONG_SYNTAX.format('BadRegex'), details) from exc
        url_pattern = parser_config.pop(Config.PROFILE_URL_PATTERN_KEY, None)
        if url_pattern is None:
            raise ProfilesError(Messages.INVALID_PROFILE.format(section), Messages.PROFILE_WITHOUT_URL)
        for parser in parsers:
            if parser_config.keys() == parser.PARAMETERS:
                break
        else:
            raise ProfilesError(Messages.INVALID_PROFILE.format(section))
        profiles[section] = Profile(url_pattern, parser, parser_config)
    return profiles


def parse_arguments(*args):
    """
    Parse each argument in args to check if it is a valid source, identify its
    type and build the corresponding handler.

    Yield tuple containing the source and its corresponding handler, which will
    be None for unsupported sources.
    """
    for arg in args:
        logging.debug(Debug.PROCESSING_ARG.format(arg))
        if is_accepted_url(arg):
            logging.debug(Debug.ARG_IS_SOURCE_SINGLE_URL)
            handler = single_url_handler(arg)
        elif arg.endswith(Config.TEXTFILE_SUFFIX):
            logging.debug(Debug.ARG_IS_SOURCE_TEXTFILE)
            handler = textfile_handler(Path(arg))
        elif arg.endswith(Config.SPREADSHEET_SUFFIX):
            logging.debug(Debug.ARG_IS_SOURCE_SPREADSHEET)
            handler = spreadsheet_handler(Path(arg))
        else:
            logging.debug(Debug.ARG_IS_SOURCE_UNSUPPORTED)
            handler = None
        yield arg, handler


# NOTE: the handlers below have to be generators for two reasons. First one, the
# list of URLs gathered by the handler from the source can be potentially large.
# So, using a generator is more efficient. Second one, this allows the caller to
# use the handler in two separate stages, first getting the URL so the handler
# keep processing it, and then sending the metadata obtained for the URL back to
# the handler. So, the handler is acting as both a coroutine and a generator.
#
# Since the metadata is sent back to the handler, it triggers another iteration
# so a double yield is needed. The first one is the one that yields the URLs and
# the second one yields a response to the caller after when the metadata is sent
# back to the handler.
#
# The first 'yield True' expression in the handlers is for signalling successful
# initialization after priming the generator/coroutine.


def single_url_handler(url):
    """
    Handle single URLs.

    The metadata for the URL is logged with INFO level, so it will be printed on
    stdout and the corresponding log files, and it is also written into a dump
    file named after the URL (properly sanitized), as key-value pairs.

    The output file has UTF-8 encoding.
    """
    sink_filename = generate_sink_filename(url_to_filename(url).with_suffix(Config.TEXTFILE_SUFFIX))
    with open(sink_filename, 'w', encoding=UTF8) as sink:
        logging.debug(Debug.DUMPING_METADATA_TO_SINK.format(sink_filename))
        yield True  # Successful initialization.
        if is_accepted_url(url):
            metadata = yield url
            yield
            if metadata:
                sink.write(Config.TEXTSINK_METADATA_HEADER.format(url))
                for key, value in metadata.items():
                    logging.debug(Debug.DUMPING_METADATA_K_V.format(key, value))
                    message = Config.TEXTSINK_METADATA_PAIR.format(key, value)
                    logging.indent()
                    logging.info(message)  # Output allowed here because it is part of the handler.
                    logging.dedent()
                    sink.write(message)
                sink.write(Config.TEXTSINK_METADATA_FOOTER)


def url_to_filename(url):
    """
    Convert the given URL to a valid filename.

    The method is quite crude but it works: replace all ASCII non-word character
    (potentially unsafe in a filename) by a character which is safe to use in a
    filename and that is visually unobtrusive so the filename is still readable.
    """
    return Path(re.sub(Config.URL_UNSAFE_CHARS_RE, Config.URL_UNSAFE_REPLACE_CHAR, url, re.ASCII))


def textfile_handler(source_filename):
    """
    Handle text files containing URLs, one per line.

    The metadata for each URL is dumped into another text file, named after the
    source file: first the URL is written, then the metadata as key-value pairs.
    Barely pretty-printed, but it is more than enough for a dump.

    All files are assumed to have UTF-8 encoding.
    """
    sink_filename = generate_sink_filename(source_filename)
    with open(source_filename, encoding=UTF8) as source:
        with open(sink_filename, 'w', encoding=UTF8) as sink:
            logging.debug(Debug.DUMPING_METADATA_TO_SINK.format(sink_filename))
            yield True  # Successful initialization.
            for url in source.readlines():
                url = url.strip()
                if not is_accepted_url(url):
                    continue
                metadata = yield url
                yield
                if metadata:
                    sink.write(Config.TEXTSINK_METADATA_HEADER.format(url))
                    for key, value in metadata.items():
                        logging.debug(Debug.DUMPING_METADATA_K_V.format(key, value))
                        sink.write(Config.TEXTSINK_METADATA_PAIR.format(key, value))
                    sink.write(Config.TEXTSINK_METADATA_FOOTER)


def spreadsheet_handler(source_filename):
    """
    Handle spreadsheets containing URLs, one per row. Ish.

    The metadata obtained for each URL is dumped into another spreadsheet, named
    after the source file, which is not created anew, it is just a copy of the
    source spreadsheet.

    The metadata is added to the spreadsheet in new columns, one per key. These
    columns are marked with a prefix as being added by the application.

    NOTE: not all sheets are processed, only the first one because it is the one
    where the URLs for the items are. Allegedly…
    """
    sink_filename = generate_sink_filename(source_filename)
    logging.debug(Debug.COPYING_WORKBOOK.format(sink_filename))

    copy2(source_filename, sink_filename)
    try:
        source_workbook = load_workbook(source_filename)
    except (KeyError, BadZipFile) as exc:
        details = str(exc).strip('"')
        details = details[0].lower() + details[1:]
        raise SourceError(Messages.INPUT_FILE_INVALID.format(type(exc).__name__)) from exc
    sink_workbook = load_workbook(sink_filename)
    yield True  # Successful initialization.

    source_sheet = source_workbook.worksheets[0]
    logging.debug(Debug.WORKING_SHEET.format(source_sheet.title))

    sink_sheet = sink_workbook.worksheets[0]
    logging.debug(Debug.INSERTING_HEADING_ROW)
    sink_sheet.insert_rows(1, 1)

    for row in source_sheet.rows:
        logging.debug(Debug.PROCESSING_ROW.format(row[0].row))
        if (url := get_url_from_row(row)) is None:
            continue
        metadata = yield url
        yield
        store_metadata_in_sheet(sink_sheet, row, metadata)
    sink_workbook.save(sink_filename)
    sink_workbook.close()
    source_workbook.close()


def get_url_from_row(row):
    """Find first URL in row."""
    url = None
    for cell in row:
        if cell.data_type != CELLTYPE_STRING:
            logging.debug(Debug.NONSTRING_CELL.format(cell.coordinate))
            continue
        if is_accepted_url(cell.value):
            url = cell.value
            logging.debug(Debug.URL_FOUND_IN_CELL.format(cell.coordinate, url))
            break  # Only the FIRST URL found in each row is considered.
    return url


def store_metadata_in_sheet(sheet, row, metadata, static = SimpleNamespace(known_metadata = {})):
    """
    Store metadata in provided sheet at given row.

    For new metadata, a new column is added to the sheet.
    For already existing metadata, the value is added to the existing column.
    """
    # NOTE: since default parameters are evaluated just once, a typical trick
    # for simulating static variables in functions is using a 'fake' default
    # parameter and using SimpleNamespace: https://stackoverflow.com/a/51437838
    if not metadata:
        return
    for key, value in metadata.items():
        key = Config.SPREADSHEET_METADATA_COLUMN_TITLE.format(key)
        if key not in static.known_metadata:
            logging.debug(Debug.NEW_METADATA_FOUND.format(key))
            column = sheet.max_column + 1
            static.known_metadata[key] = column
            logging.debug(Debug.METADATA_STORED_IN_COLUMN.format(key, get_column_letter(column)))
            cell = sheet.cell(row=1, column=column, value=key)
            cell.font = Font(name=Config.SPREADSHEET_CELL_FONT)
            cell.fill = PatternFill(start_color=Config.SPREADSHEET_CELL_COLOR, fill_type=Config.SPREADSHEET_CELL_FILL)
                # Set column width.
                #
                # As per Excel specification, the width units are the width of
                # the zero character of the font used by the Normal style for a
                # workbook. So a column of width 10 would fit exactly 10 zero
                # characters in the font specified by the Normal style.
                #
                # No, no kidding.
                #
                # Since this width units are, IMHO, totally arbitrary, let's
                # choose an arbitrary column width. To wit, the Answer to the
                # Ultimate Question of Life, the Universe, and Everything.
            sheet.column_dimensions[get_column_letter(column)].width = 42
                # This is needed because sometimes Excel files are not properly
                # generated and the last column has a 'max' field too large, and
                # that has an unintended consequence: ANY change to the settings
                # of that column affects ALL the following ones whose index is
                # less than 'max'… So, it's better to fix that field.
            sheet.column_dimensions[get_column_letter(column)].max = column
        logging.debug(Debug.DUMPING_METADATA_K_V.format(key, value))
            # Since a heading row is inserted, the rows where metadata has to go
            # have now an +1 offset, as they have been displaced.
        sheet.cell(row[0].row + 1, static.known_metadata[key], value=value)


def bootstrap(handler):
    """Bootstrap (prime) and handle initialization errors for handler."""
    if handler is None:
        raise SourceError(Messages.UNSUPPORTED_SOURCE)

    try:
        handler.send(None)
    except FileNotFoundError as exc:
        raise SourceError(Messages.INPUT_FILE_NOT_FOUND) from exc
    except PermissionError as exc:
        if Path(exc.filename).stem.endswith(Config.SINKFILE_STEM):
            raise SourceError(Messages.OUTPUT_FILE_NO_PERMISSION) from exc
        raise SourceError(Messages.INPUT_FILE_NO_PERMISSION) from exc


def get_parser(url, profiles):
    """
    Return the appropriate parser for the url.

    The appropriate parser is retrieved by finding the profile within profiles
    whose url_pattern matches the given url. Then, the parser is returned after
    being properly configured to handle the url contents.
    """
    for profile_name, profile in profiles.items():
        if profile.url_pattern.search(url):
            logging.debug(Debug.DETECTED_PROFILE.format(profile_name))
            profile.parser.configure(profile.parser_config)
            return profile.parser
    return None


def saca_las_mantecas(url, parser):
    """
    Saca las mantecas from the provided url, that is, retrieve its contents,
    parse them using parser, and obtain library catalogue metadata, if any.

    Return obtained metadata as a dictionary.
    """
    try:
        contents, encoding = retrieve_url(url)
    except URLError as exc:
        # Depending on the particular error which happened, the reason attribute
        # of the URLError exception can be a simple error message, an instance
        # of some OSError derived Exception, etc. So, in order to have useful
        # messages, some discrimination has to be done here.
        if isinstance(exc.reason, OSError):
            try:
                error_code = errno.errorcode[exc.reason.errno]
            except KeyError:
                error_code = exc.reason.errno
            except AttributeError:
                error_code = Messages.UNKNOWN_ERRNO
            error_reason = exc.reason.strerror
            details = Messages.OSLIKE_URLERROR
        elif isinstance(exc, HTTPError):
            error_code = exc.code
            error_reason = exc.reason.lower()
            details = Messages.HTTP_PROTOCOL_URLERROR
        else:
            error_code = EMPTY_STRING
            error_reason = exc.reason
            details = Messages.GENERIC_URLERROR
        error_reason = (error_reason[0].lower() + error_reason[1:]).rstrip('.')
        raise SkimmingError(Messages.URL_ACCESS_ERROR, details.format(error_code, error_reason)) from exc
    # Apparently, HTTPException, ConnectionError and derived exceptions are
    # masked or wrapped by urllib, and documentation is not very informative.
    # So, just in case something weird happen, it is better to handle these
    # exception types as well.
    except HTTPException as exc:
        details = f'{type(exc).__name__}: {exc}.'
        raise SkimmingError(Messages.HTTP_RETRIEVAL_ERROR, details) from exc
    except ConnectionError as exc:
        try:
            error_code = errno.errorcode[exc.errno]
        except (AttributeError, KeyError):
            error_code = Messages.UNKNOWN_ERRNO
        details = f'{exc.strerror.capitalize().rstrip(".")}.'
        raise SkimmingError(Messages.CONNECTION_ERROR.format(error_code), details) from exc

    if not contents:
        raise SkimmingError(Messages.NO_CONTENTS_ERROR)

    parser.feed(contents.decode(encoding))
    parser.close()
    if metadata := parser.get_metadata():
        return metadata
    raise SkimmingError(Messages.NO_METADATA_FOUND)


def retrieve_url(url):
    """
    Retrieve contents from url.

    First resolve any meta http-equiv="Refresh" redirection for url and then get
    the contents as a byte string.

    Then, detect the contents encoding (in HTML jargon, the charset).

    Return a (contents, charset) tuple.
    """
    if not is_accepted_url(url):
        raise URLError(Messages.UNKNOWN_URL_TYPE.format(url))

    if url.startswith('file://'):
        url = resolve_file_url(url)

    while url:
        logging.debug(Debug.PROCESSING_URL.format(url))
        with urlopen(Request(url, headers={'User-Agent': Config.USER_AGENT})) as response:
            # First, check if any redirection is needed and get the charset the easy way.
            contents = response.read()
            charset = response.headers.get_content_charset()
        url = get_redirected_url(url, contents)

    # In this point, we have the contents as a byte string.
    # If the charset is None, it has to be determined the hard way.
    if charset is None:
        logging.debug(Debug.CHARSET_NOT_IN_HEADERS)
        charset = detect_html_charset(contents)
    else:
        logging.debug(Debug.CHARSET_IN_HEADERS)
    logging.debug(Debug.CONTENTS_ENCODING.format(charset))

    return contents, charset


def resolve_file_url(url):
    """Resolve relative paths in file: url."""
    parsed_url = urlparse(url)
    resolved_path = unquote(parsed_url.path[1:])
    resolved_path = Path(resolved_path).resolve().as_posix()
    resolved_path = quote(resolved_path, safe=Config.FILE_URL_SAFE_CHARS)
    return parsed_url._replace(path=resolved_path).geturl()


def get_redirected_url(base_url, contents):
    """
    Get redirected URL from a meta http-equiv="refresh" pragma in contents. Use
    base_url as base URL for redirection, if some parts are missing in the URL
    specified by the pragma.

    Return redirected URL, or None if there is no redirection pragma.
    """
    if match := re.search(Config.META_REFRESH_RE, contents, re.I):
        base_url = urlparse(base_url)
        redirected_url = urlparse(match.group(1).decode(ASCII))
        for field in base_url._fields:
            value = getattr(base_url, field)
            # If not specified in the redirected URL, both the scheme and netloc
            # will be reused from the base URL. Any other field will be obtained
            # from the redirected URL and used, no matter if it is empty.
            if field in ('scheme', 'netloc') and not getattr(redirected_url, field):
                redirected_url = redirected_url._replace(**{field: value})
        redirected_url = urlunparse(redirected_url)
        logging.debug(Debug.REDIRECTED_URL.format(redirected_url))
        return redirected_url
    return None


def detect_html_charset(contents):
    """
    Detect contents charset from HTML tags, if any, and return it.

    If the charset can not be determined, a sane fallback is used. It may look
    like UTF-8 would be such a sane fallback, because modern web pages may NOT
    specify any encoding if they are using UTF-8 and it is identical to ASCII
    for 7-bit codepoints, but the problem is that UTF-8 will fail for web pages
    whose encoding is any ISO/IEC 8859 variant.

    So, the sane default is another, set in the global configuration, and it is
    based on the encoding most frequently used by the web pages this application
    will generally process.
    """
    charset = Config.FALLBACK_HTML_CHARSET
    if match := re.search(Config.META_HTTP_EQUIV_CHARSET_RE, contents, re.I):
        # Next best thing, from the meta http-equiv="content-type".
        logging.debug(Debug.CHARSET_FROM_HTTP_EQUIV)
        charset = match.group(1).decode(ASCII)
    elif match := re.search(Config.META_CHARSET_RE, contents, re.I):
        # Last resort, from some meta charset, if any…
        logging.debug(Debug.CHARSET_FROM_META_CHARSET)
        charset = match.group(1).decode(ASCII)
    else:
        logging.debug(Debug.CHARSET_FROM_DEFAULT)
    return charset


@loggerize
@keyboard_interrupt_handler
def main(*args):
    """."""
    exitcode = ExitCodes.SUCCESS

    if not args:
        # Input arguments should be provided automatically to the application if
        # it is used as a drag'n'drop target which is actually the intended way
        # of operation, generally speaking.
        #
        # But the application can be also run by hand from a command prompt, so
        # it is better to signal the end user with an error and explanation if
        # no input arguments are provided, as soon as possible.
        error(Messages.NO_ARGUMENTS)
        return ExitCodes.NO_ARGUMENTS

    try:
        profiles = load_profiles(Config.INIFILE_PATH)
        if not profiles:
            raise ProfilesError(Messages.EMPTY_PROFILES.format(Config.INIFILE_PATH))
        logging.debug(Debug.FOUND_PROFILES.format(', '.join(profiles.keys())))
    except ProfilesError as exc:
        error(exc, exc.details)
        return ExitCodes.ERROR

    logging.info(Messages.SKIMMING_MARKER)
    logging.indent()
    for source, handler in parse_arguments(*args):
        logging.info(Messages.SOURCE_LABEL.format(source))
        try:
            bootstrap(handler)
        except SourceError as exc:
            logging.indent()
            warning(exc)
            logging.dedent()
            exitcode = ExitCodes.WARNING
            continue

        logging.indent()
        for url in handler:
            logging.info(url)
            metadata = {}
            try:
                if (parser := get_parser(url, profiles)) is None:
                    raise SkimmingError(Messages.NO_MATCHING_PROFILE)
                metadata = saca_las_mantecas(url, parser)
            except SkimmingError as exc:
                logging.indent()
                warning(exc)
                logging.debug(exc.details)
                logging.dedent()
                exitcode = ExitCodes.WARNING
            finally:
                # No matter if metadata has actual contents or not, the handler
                # has to be 'advanced' to the next URL, so the metadata it is
                # expecting has to be sent to the handler.
                handler.send(metadata)
        logging.dedent()
    logging.dedent()
    return exitcode


if __name__ == '__main__':
    atexit.register(wait_for_keypress)
    sys.excepthook = excepthook
    sys.exit(main(*sys.argv[1:]))
