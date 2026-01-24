#! /usr/bin/env python
"""See 'README.md' for details."""
import sys  # noqa: I001
if sys.platform != 'win32':
    sys.stdout.write('\nThis program is compatible only with the Win32 platform.\n')
    sys.stdout.flush()
    sys.exit(None)

# pylint: disable=wrong-import-position
import atexit
from collections.abc import Callable, Generator
import configparser
import contextlib
from ctypes import byref, c_uint, create_unicode_buffer, windll
from ctypes.wintypes import MAX_PATH as MAX_PATH_LEN
from enum import auto, IntEnum, StrEnum
import errno
from functools import wraps
from html.parser import HTMLParser
from http.client import HTTPException
from importlib.metadata import version
import logging
from logging.config import dictConfig
from msvcrt import get_osfhandle, getch
from pathlib import Path
import platform
import re
from shutil import copy2
import time
import traceback as tb
from types import SimpleNamespace, TracebackType
from typing import Any, cast, ClassVar, LiteralString, NamedTuple, TYPE_CHECKING
from urllib.error import HTTPError, URLError
from urllib.parse import quote, unquote, urlparse, urlunparse
from urllib.request import Request, urlopen
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell, TYPE_STRING as CELLTYPE_STRING
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter

if TYPE_CHECKING:
    from io import TextIOWrapper

    from openpyxl.worksheet.worksheet import Worksheet

# Handlers are not implemented as classes, but as generators.
type Handler = Generator[str, dict[str, str] | None]


class Constants:  # pylint: disable=too-few-public-methods
    """Program configuration values."""

    APP_PATH = Path(__file__)
    APP_NAME = APP_PATH.stem
    APP_VERSION = version(APP_NAME)
    APP_REPOSITORY = 'https://github.com/DervishD/sacamantecas'
    APP_PLATFORM = f'Windows {platform.version()};{platform.architecture()[0]};{platform.machine()}'

    DEVELOPMENT_MODE = 'dev' in APP_VERSION

    UTF8 = 'utf-8'
    ASCII = 'ascii'
    DOUBLE_QUOTE_CHAR = '"'
    PERIOD = '.'

    OUTPUT_SEPARATOR = ', '

    ERROR_MARKER = '*** '
    ERROR_PAYLOAD_INDENT = len(ERROR_MARKER)

    TIMESTAMP_FORMAT = '%Y%m%d_%H%M%S'

    USER_AGENT = f'{APP_NAME}/{APP_VERSION} +{APP_REPOSITORY} ({APP_PLATFORM})'

    ACCEPTED_URL_SCHEMES = ('https', 'http', 'file')

    FALLBACK_HTML_CHARSET = 'ISO-8859-1'

    PROFILE_URL_PATTERN_KEY = 'url'
    PROFILE_BAD_REGEX_ERROR = 'BadRegex'

    TEXTFILE_SUFFIX = '.txt'
    SPREADSHEET_SUFFIX = '.xlsx'

    TIMESTAMP_STEM = time.strftime(f'_{TIMESTAMP_FORMAT}')
    SINKFILE_STEM = '_out'

    ROOT_PATH = Path(sys.executable if getattr(sys, 'frozen', False) else __file__).resolve().parent

    MAIN_OUTPUT_PATH = ROOT_PATH / f'{APP_NAME}_log{"" if DEVELOPMENT_MODE else TIMESTAMP_STEM}{TEXTFILE_SUFFIX}'
    FULL_OUTPUT_PATH = ROOT_PATH / f'{APP_NAME}_trace{"" if DEVELOPMENT_MODE else TIMESTAMP_STEM}{TEXTFILE_SUFFIX}'
    INIFILE_PATH = ROOT_PATH / f'{APP_NAME}.ini'

    LOGGING_INDENTCHAR = ' '
    LOGGING_FORMAT_STYLE = '{'
    LOGGING_LEVELNAME_MAX_LEN = len(max(logging.getLevelNamesMapping(), key=len))
    LOGGING_LEVELNAME_SEPARATOR = '| '
    LOGGING_LONG_FORMAT = (
        '{asctime}.{msecs:04.0f} '
        f'{{levelname:{LOGGING_LEVELNAME_MAX_LEN}}}'
        f'{LOGGING_LEVELNAME_SEPARATOR}'
        '{message}'
    )
    LOGGING_FALLBACK_FORMAT = '{message}'
    LOGGING_SHORT_FORMAT = '{asctime} {message}'
    LOGGING_CONSOLE_FORMAT = '{message}'

    HANDLER_BOOTSTRAP_SUCCESS = 'Handler bootstrap successful.'

    TEXTSINK_METADATA_HEADER = '{}\n'
    TEXTSINK_METADATA_INDENT = '  '
    TEXTSINK_METADATA_SEPARATOR = ': '
    TEXTSINK_METADATA_PAIR = f'{TEXTSINK_METADATA_INDENT}{{}}{TEXTSINK_METADATA_SEPARATOR}{{}}\n'
    TEXTSINK_METADATA_FOOTER = '\n'

    METADATA_KEY_TERMINATOR = ':'

    SPREADSHEET_METADATA_COLUMN_MARKER = '[sm] '
    SPREADSHEET_METADATA_COLUMN_TITLE = f'{SPREADSHEET_METADATA_COLUMN_MARKER}{{}}'
    SPREADSHEET_CELL_FONT = 'Calibri'
    SPREADSHEET_CELL_COLOR = 'baddad'
    SPREADSHEET_CELL_FILL = 'solid'

    URL_UNSAFE_CHARS_RE = r'\W'
    URL_UNSAFE_REPLACE_CHAR = '_'
    FILE_SCHEME = 'file://'
    FILE_URL_SAFE_CHARS = ':/'
    FILE_URL_SEPARATOR = '/'

    USER_AGENT_HEADER = 'User-Agent'

    META_HTTP_EQUIV_CHARSET_RE = rb'<meta http-equiv="content-type".*charset="([^"]+)"'
    META_CHARSET_RE = rb'<meta charset="([^"]+)"'
    META_REFRESH_RE = rb'<meta http-equiv="refresh" content="(?:[^;]+;\s+)?URL=([^"]+)"'
    URL_FIELDS_TO_REUSE = ('scheme', 'netloc')


class Messages(StrEnum):
    """Messages."""

    PRESS_ANY_KEY = '\nPulse cualquier tecla para continuar...'
    KEYBOARD_INTERRUPT = 'El usuario interrumpió la operación de la aplicación.'
    NO_ARGUMENTS = (
        'No se han especificado fuentes de entrada para ser procesadas.\n'
        '\n'
        'Arrastre y suelte un fichero de entrada sobre el icono de la aplicación,\n'
        'o bien proporcione los nombres de las fuentes de entrada como argumentos.'
    )

    DEBUGGING_INIT = 'Registro de depuración iniciado.'
    APP_BANNER = f'{Constants.APP_NAME} versión {Constants.APP_VERSION}'
    PROCESS_DONE = '\nProceso finalizado.'
    DEBUGGING_DONE = 'Registro de depuración finalizado.'

    ERROR_HEADER = f'\n{Constants.ERROR_MARKER}Error en {Constants.APP_NAME}.\n'
    WARNING_HEADER = '* Aviso: '
    ERROR_DETAILS_HEADING = '\nInformación adicional sobre el error:'
    ERROR_DETAILS_PREAMBLE = '│ '
    ERROR_DETAILS_TAIL = '╰'

    UNEXPECTED_OSERROR = 'Error inesperado del sistema operativo.'
    OSERROR_DETAILS = (
        '     type = {}\n'
        '    errno = {}\n'
        ' winerror = {}\n'
        ' strerror = {}\n'
        ' filename = {}\n'
        'filename2 = {}\n'
    )
    OSERROR_DETAIL_NA = '[No disponible]'
    UNHANDLED_EXCEPTION = 'Excepción sin gestionar.'
    EXCEPTION_DETAILS = 'type = {}\nvalue = {}\nargs: {}'
    EXCEPTION_DETAILS_ARG = '\n  [{}] {}'
    TRACEBACK_HEADER = '\n\ntraceback:\n{}'
    TRACEBACK_FRAME_HEADER = '▸ {}\n'
    TRACEBACK_FRAME_LINE = '  {}, {}: {}\n'
    TRACEBACK_TOPLEVEL_FRAME = '<module>'
    UNKNOWN_ERRNO = 'desconocido'

    PROCESSING_ARG = 'Procesando argumento «{}».'
    ARG_IS_SOURCE_SINGLE_URL = 'El argumento es una fuente de tipo single_url.'
    ARG_IS_SOURCE_TEXTFILE = 'El argumento es una fuente de tipo textfile.'
    ARG_IS_SOURCE_SPREADSHEET = 'El argumento es una fuente de tipo spreadsheet.'
    ARG_IS_SOURCE_UNSUPPORTED = 'El argumento no es un tipo de fuente admitido.'

    LOADING_PROFILES = 'Obteniendo perfiles desde «{}».'
    EMPTY_PROFILES = 'No hay perfiles definidos en el fichero de perfiles «{}».'
    MISSING_PROFILES = 'No se encontró o no se pudo leer el fichero de perfiles «{}».'
    PROFILES_WRONG_SYNTAX = 'Error de sintaxis «{}» leyendo el fichero de perfiles.'
    PROFILES_WRONG_SYNTAX_DETAILS = 'Perfil «{}», {}:\n  {}{}{}\n  {:_<{}}^'
    PROFILES_WRONG_SYNTAX_DETAILS_SEPARATOR = ' = '
    INVALID_PROFILE = 'El perfil «{}» es inválido.'
    PROFILE_WITHOUT_URL = 'El perfil no incluye un patrón de URL.'
    UNKNOWN_URL_TYPE = 'El URL «{}» es de tipo desconocido.'
    NO_MATCHING_PROFILE = 'No se encontró un perfil para procesar el URL.'
    FOUND_PROFILES = 'Se obtuvieron los siguientes perfiles: {}.'
    DETECTED_PROFILE = 'Perfil detectado: «{}».'

    PROCESSING_URL = 'Procesando URL «{}».'
    REDIRECTED_URL = 'URL redirigido a «{}».'
    OSLIKE_URLERROR = 'Error de red {}: {}.'
    HTTP_PROTOCOL_URLERROR = 'Error de protocolo HTTP {}: {}.'
    GENERIC_URLERROR = 'Error de URL{}: {}.'
    URL_ACCESS_ERROR = 'No resultó posible acceder a la dirección especificada.'
    HTTP_RETRIEVAL_ERROR = 'No se obtuvieron contenidos.'
    CONNECTION_ERROR = 'Se produjo un error de conexión «{}» accediendo al URL.'
    NO_CONTENTS_ERROR = 'No se recibieron contenidos del URL.'

    CHARSET_NOT_IN_HEADERS = 'Charset no detectado en las cabeceras.'
    CHARSET_IN_HEADERS = 'Charset detectado en las cabeceras.'
    CHARSET_FROM_HTTP_EQUIV = 'Charset detectado mediante meta http-equiv.'
    CHARSET_FROM_META_CHARSET = 'Charset detectado mediante meta charset.'
    CHARSET_FROM_DEFAULT = 'Charset not detectado, usando valor por defecto.'
    CONTENTS_ENCODING = 'Contenidos codificados con charset «{}».'

    SKIMMING_MARKER = '\nSacando las mantecas:\n'
    SOURCE_LABEL = 'Fuente: {}'
    UNSUPPORTED_SOURCE = 'La fuente no es de un tipo admitido.'
    INPUT_FILE_NOT_FOUND = 'No se encontró el fichero de entrada.'
    INPUT_FILE_NO_PERMISSION = 'No hay permisos suficientes para leer el fichero de entrada.'
    OUTPUT_FILE_NO_PERMISSION = 'No hay permisos suficientes para crear el fichero de salida.'
    NO_METADATA_FOUND = 'No se obtuvieron metadatos.'

    HTML_START_TAG = '➜ HTML <{}{}{}>'
    METADATA_KEY_FOUND = 'Se encontró la clave «{}».'
    METADATA_VALUE_FOUND = 'Se encontró el valor «{}».'
    METADATA_IS_EMPTY = 'Metadato vacío.'
    METADATA_MISSING_VALUE = 'Metadato «{}» incompleto, ignorando.'
    METADATA_MISSING_KEY = 'No se encontró una clave, usando «{}».'
    METADATA_OK = 'Metadato correcto «{}: {}».'
    METADATA_KEY_MARKER_FOUND = 'Se encontró una marca de clave «{}».'
    METADATA_VALUE_MARKER_FOUND = 'Se encontró una marca de valor «{}».'
    METADATA_MARKER_FOUND = 'Se encontró una marca de metadato «{}».'
    DUMPING_METADATA_TO_SINK = 'Volcando metadatos a «{}».'
    DUMPING_METADATA_K_V = 'Añadiendo metadato «{}» con valor «{}».'
    PARSER_NESTING_ERROR_K_IN_V = 'Problema de anidación (clave dentro de valor), restableciendo parser.'
    PARSER_NESTING_ERROR_V_IN_K = 'Problema de anidación (valor dentro de clave), restableciendo parser.'

    COPYING_WORKBOOK = 'Copiando workbook a «{}».'
    WORKING_SHEET = 'La hoja con la que se trabajará es «{}»".'
    SOURCE_SHEET_IS_INVALID = 'La hoja de entrada es inválida ({}).'
    INSERTING_HEADING_ROW = 'Insertando fila de cabeceras.'
    PROCESSING_ROW = 'Procesando fila {}.'
    NONSTRING_CELL = 'La celda «{}» no es de tipo cadena, será ignorada.'
    URL_FOUND_IN_CELL = 'Se encontró un URL en la celda «{}»: {}'
    NEW_METADATA_FOUND = 'Se encontró un metadato nuevo, «{}».'
    METADATA_STORED_IN_COLUMN = 'El metadato «{}» irá en la columna «{}».'


class ExitCodes(IntEnum):
    """Standardized exit codes."""  # noqa: D204
    SUCCESS = 0
    NO_ARGUMENTS = 1
    WARNING = 2
    ERROR = 3
    KEYBOARD_INTERRUPT = 127


# Needed for having VERY basic logging when the code is imported rather
# than run (for example, when running tests).
logging.basicConfig(
    level=logging.NOTSET,
    style=Constants.LOGGING_FORMAT_STYLE,
    format=Constants.LOGGING_FALLBACK_FORMAT,
    force=True,
)


# Reconfigure standard output streams so they use UTF-8 encoding even if
# they are redirected to a file when running the program from a shell.
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    cast('TextIOWrapper', sys.stdout).reconfigure(encoding=Constants.UTF8)
if sys.stderr and hasattr(sys.stdout, 'reconfigure'):
    cast('TextIOWrapper', sys.stderr).reconfigure(encoding=Constants.UTF8)


class CustomLogger(logging.Logger):
    """Custom logger with indentation support."""

    INCREASE_INDENT_SYMBOL = '+'
    DECREASE_INDENT_SYMBOL = '-'

    def __init__(self, name: str, level: int = logging.NOTSET) -> None:
        """Initialize logger with a *name* and an optional *level*."""
        super().__init__(name, level)
        self.indentlevel: int = 0
        self.indentation = ''

    def makeRecord(self, *args: Any, **kwargs: Any) -> logging.LogRecord:  # noqa: ANN401, N802
        """Create a new logging record with indentation support."""
        record = super().makeRecord(*args, **kwargs)
        record.msg = '\n'.join(f'{self.indentation}{line}'.rstrip() for line in record.msg.split('\n'))
        return record

    def _set_indentlevel(self, level: int | LiteralString) -> None:
        """Set current logging indentation to *level*.

        If *level* is:
            - `INCREASE_INDENT_SYMBOL` string, indentation is increased
            - `DECREASE_INDENT_SYMBOL` string, indentation is decreased
            - Any integer `>=0`, indentation is set to that value

        For any other value, `ValueError` is raised.

        Not for public usage, use `self.set_indent(level)` instead.
        """
        if level == self.INCREASE_INDENT_SYMBOL:
            self.indentlevel += 1
        if level == self.DECREASE_INDENT_SYMBOL:
            self.indentlevel = max(0, self.indentlevel - 1)
        if isinstance(level, int) and level >= 0:
            self.indentlevel = level
        self.indentation = Constants.LOGGING_INDENTCHAR * self.indentlevel

    def set_indent(self, level: int) -> None:
        """Set current logging indentation to *level*.

        *level* can be any positive integer or zero.

        For any other value, `ValueError` is raised.
        """
        self._set_indentlevel(max(0, level))

    def indent(self) -> None:
        """Increment current logging indentation level."""
        self._set_indentlevel(self.INCREASE_INDENT_SYMBOL)

    def dedent(self) -> None:
        """Decrement current logging indentation level."""
        self._set_indentlevel(self.DECREASE_INDENT_SYMBOL)

    def config(self, *, main_log_output: Path, full_log_output: Path) -> None:
        """Configure logger.

        With the default configuration **ALL** logging messages are sent
        to *full_log_output* using a detailed format which includes the
        current timestamp and some debugging information; messages with
        severity of `logging.INFO` or higher, intended to be the typical
        program output, are sent to *main_log_output*, also timestamped.
        """
        class MultilineFormatter(logging.Formatter):
            """Simple custom formatter with multiline support."""  # noqa: D204
            def format(self, record: logging.LogRecord) -> str:
                """Format multiline records so they look like multiple records."""
                formatted_record = super().format(record)
                preamble = formatted_record[0:formatted_record.rfind(record.message)]
                return '\n'.join(f'{preamble}{line}'.rstrip() for line in record.message.split('\n'))

        logging_configuration: dict[str, Any] = {
            'version': 1,
            'disable_existing_loggers': False,
            'loggers': {
                self.name: {
                    'level': logging.NOTSET,
                    'propagate': False,
                    'handlers': [],
                },
            },
        }

        formatters = {}
        handlers = {}

        if full_log_output:
            formatters['full_log_formatter'] = {
                '()': MultilineFormatter,
                'style': Constants.LOGGING_FORMAT_STYLE,
                'format': Constants.LOGGING_LONG_FORMAT,
                'datefmt': Constants.TIMESTAMP_FORMAT,
            }
            handlers['full_log_handler'] = {
                'level': logging.NOTSET,
                'formatter': 'full_log_formatter',
                'class': logging.FileHandler,
                'filename': full_log_output,
                'mode': 'w',
                'encoding': Constants.UTF8,
            }

        if main_log_output:
            formatters['main_log_formatter'] = {
                '()': MultilineFormatter,
                'style': Constants.LOGGING_FORMAT_STYLE,
                'format': Constants.LOGGING_SHORT_FORMAT,
                'datefmt': Constants.TIMESTAMP_FORMAT,
            }
            handlers['main_log_handler'] = {
                'level': logging.INFO,
                'formatter': 'main_log_formatter',
                'class': logging.FileHandler,
                'filename': main_log_output,
                'mode': 'w',
                'encoding': Constants.UTF8,
            }

        formatters['console_formatter'] = {
            '()': MultilineFormatter,
            'style': Constants.LOGGING_FORMAT_STYLE,
            'format': Constants.LOGGING_CONSOLE_FORMAT,
        }
        handlers['stdout_handler'] = {
            'level': logging.NOTSET,
            'formatter': 'console_formatter',
            'filters': [lambda record: (record.levelno == logging.INFO)],  # type: ignore  # noqa: PGH003
            'class': logging.StreamHandler,
            'stream': sys.stdout,
        }
        handlers['stderr_handler'] = {
            'level': logging.WARNING,
            'formatter': 'console_formatter',
            'class': logging.StreamHandler,
            'stream': sys.stderr,
        }

        logging_configuration['formatters'] = formatters
        logging_configuration['handlers'] = handlers
        logging_configuration['loggers'][self.name]['handlers'] = handlers.keys()
        dictConfig(logging_configuration)
logging.setLoggerClass(CustomLogger)
logger: CustomLogger = cast('CustomLogger', logging.getLogger(Constants.APP_NAME))


class BaseCustomError(Exception):
    """Base class for all program custom exceptions."""  # noqa: D204
    # pylint: disable-next=keyword-arg-before-vararg
    def __init__ (self, message: str, details: object = '', *args: object, **kwargs: object) -> None:
        """Initialize exception with message and details."""
        self.details = details
        super().__init__(message, *args, **kwargs)

class ProfilesError(BaseCustomError):
    """Raise for profile-related errors."""

class SourceError(BaseCustomError):
    """Raise for source-related errors."""

class SkimmingError(BaseCustomError):
    """Raise for skimming-related errors."""


class BaseParser(HTMLParser):
    """Base class for catalogue parsers."""

    PARAMETERS: ClassVar[set[str]] = set()
    DEFAULT_K = ''
    DEFAULT_V = ''
    EMPTY_KEY_PLACEHOLDER = '[vacío]'
    MULTIDATA_SEPARATOR = ' / '
    MULTIVALUE_SEPARATOR = ' === '

    def __init__(self, *args: object, **kwargs: object) -> None:
        """Initialize object."""
        self.within_k: bool
        self.within_v: bool
        self.current_k: str
        self.current_v: str
        self.last_k: str
        self.retrieved_metadata: dict[str, list[str]]
        self.config: dict[str, re.Pattern[str]]
        super().__init__(*args, **kwargs)

    def reset(self) -> None:
        """Reset parser state. Called implicitly from `__init__()`."""
        super().reset()
        self.within_k = self.within_v = False
        self.current_k = self.DEFAULT_K
        self.current_v = self.DEFAULT_V
        self.last_k = self.DEFAULT_K
        self.retrieved_metadata = {}

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        """Handle opening tags."""
        logger.debug(Messages.HTML_START_TAG.format(tag, ' ' * bool(attrs), ' '.join((f'{k}="{v}"' for k, v in attrs))))

    def handle_data(self, data: str) -> None:
        """Handle data."""
        if self.within_k or self.within_v:
            # Clean up received data by removing superfluous whitespace
            # characters, including newlines, carriage returns, etc.
            data = ' '.join(data.split())
            if not data:
                return
        if self.within_k:
            logger.debug(Messages.METADATA_KEY_FOUND.format(data))
            self.current_k += data.rstrip(Constants.METADATA_KEY_TERMINATOR)
            self.last_k = self.current_k
            return
        if self.within_v:
            logger.debug(Messages.METADATA_VALUE_FOUND.format(data))
            self.current_v += f'{self.MULTIDATA_SEPARATOR if self.current_v else ''}{data}'
            return

    def configure(self, config: dict[str, re.Pattern[str]]) -> None:
        """Configure parser.

        Set up parser with configuration from *config*, that is, a set
        of values for the supported parser parameters.

        Only supported configuration parameters are used, the rest are
        silently ignored.

        This operation also resets the parser.
        """
        self.reset()
        self.config = {key: value for key, value in config.items() if key in self.PARAMETERS}

    def store_metadata(self) -> None:
        """Store found metadata, handling missing parts."""
        if not self.current_k and not self.current_v:
            logger.debug(Messages.METADATA_IS_EMPTY)
        if self.current_k and not self.current_v:
            logger.debug(Messages.METADATA_MISSING_VALUE.format(self.current_k))
        if not self.current_k and self.current_v:
            self.current_k = self.last_k if self.last_k else self.EMPTY_KEY_PLACEHOLDER
            logger.debug(Messages.METADATA_MISSING_KEY.format(self.current_k))
        if self.current_k and self.current_v:
            if self.current_k not in self.retrieved_metadata:
                self.retrieved_metadata[self.current_k] = []
            # A set is not used below, to preserve order.
            if self.current_v not in self.retrieved_metadata[self.current_k]:
                self.retrieved_metadata[self.current_k].append(self.current_v)
            logger.debug(Messages.METADATA_OK.format(self.current_k, self.current_v))
        self.current_k = self.DEFAULT_K
        self.current_v = self.DEFAULT_V

    def get_metadata(self) -> dict[str, str]:
        """Get the metadata retrieved so far."""
        metadata: dict[str, str] = {}
        for key, value in self.retrieved_metadata.items():
            metadata[key] = self.MULTIVALUE_SEPARATOR.join(value)
        return metadata


class OldRegimeParser(BaseParser):  # pylint: disable=unused-variable
    """Parser for *Old Regime* catalogues.

    Parser for *Old Regime* catalogues, which use different HTML `class`
    attributes to mark metadata keys and metadata values.

    This is inherently complex, specially when dealing with ill-formed
    HTML, so in order to keep this parser as simple as possible, some
    assumptions are made.

    See the comments below to know which those assumptions are.
    """

    K_CLASS = 'k_class'
    V_CLASS = 'v_class'
    PARAMETERS = BaseParser.PARAMETERS | {K_CLASS, V_CLASS}
    CLASS_ATTR = 'class'

    def __init__(self, *args: object, **kwargs: object) -> None:
        """Initialize object."""
        self.current_k_tag = None
        self.current_v_tag = None
        super().__init__(*args, **kwargs)

    def reset(self) -> None:
        """Reset parser state. Called implicitly from `__init__()`."""
        super().reset()
        self.current_k_tag = self.current_v_tag = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        """Handle opening tags."""
        super().handle_starttag(tag, attrs)
        for attr in attrs:
            if attr[1] is None:
                continue
            if attr[0] == self.CLASS_ATTR and (match := self.config[self.K_CLASS].search(attr[1])):
                logger.debug(Messages.METADATA_KEY_MARKER_FOUND.format(match.group(0)))
                self.within_k = True
                self.current_k = self.DEFAULT_K
                self.current_k_tag = tag
                if self.within_v:
                    # If still processing a value, log about the nesting
                    # error but reset parser so everything starts afresh
                    # like if a new key had been found.
                    logger.debug(Messages.PARSER_NESTING_ERROR_K_IN_V)
                    self.within_v = False
                    self.current_v = self.DEFAULT_V
                    self.current_v_tag = None
                break
            if attr[0] == self.CLASS_ATTR and (match := self.config[self.V_CLASS].search(attr[1])):
                logger.debug(Messages.METADATA_VALUE_MARKER_FOUND.format(match.group(0)))
                self.within_v = True
                self.current_v = self.DEFAULT_V
                self.current_v_tag = tag
                if self.within_k:
                    # If a new key is found but a previous one is still
                    # being processed, this nesting error is not fatal
                    # and can still be recovered, up to a certain point.
                    # If some data was got for the key, the parser will
                    # be left in 'within_v' mode to try to get the value
                    # if possible. Otherwise the parser is reset.
                    logger.debug(Messages.PARSER_NESTING_ERROR_V_IN_K)
                    self.within_k = False
                    self.current_k_tag = None
                    if not self.current_k:
                        self.within_v = False
                        self.current_v_tag = None
                break

    def handle_endtag(self, tag: str) -> None:
        """Handle closing tags."""
        super().handle_endtag(tag)
        if self.within_k and tag == self.current_k_tag:
            self.within_k = False
            self.current_k_tag = None
            return
        if self.within_v and tag == self.current_v_tag:
            self.within_v = False
            self.current_v_tag = None
            self.store_metadata()
            return


class BaratzParser(BaseParser):   # pylint: disable=unused-variable
    """Parser for *Baratz* catalogues.

    Parser for catalogues whose contents have been generated by the new
    *Baratz* frontend, which does not use the HTML `class` attributes to
    mark metadata keys and values. Instead a particular HTML element is
    used as a kind of marker for the start of metadata blocks, which in
    turn contain the metadata as a list of `<dt>/<dd>` pairs.

    Within that list, the `<dt>` HTML element contains the metadata key,
    whereas the `<dd>` HTML element containing the metadata value.

    The element which acts as the marker for a metadata block is defined
    by its HTML tag, the attribute which contains the marker value, and
    of course the marker value itself.
    """

    M_TAG = 'm_tag'
    M_ATTR = 'm_attr'
    M_VALUE = 'm_value'
    PARAMETERS = BaseParser.PARAMETERS | {M_TAG, M_ATTR, M_VALUE}
    K_TAG = 'dt'
    V_TAG = 'dd'

    def __init__(self, *args: object, **kwargs: object) -> None:
        """Initialize object."""
        self.within_meta: bool
        super().__init__(*args, **kwargs)

    def reset(self) -> None:
        """Reset parser state. Called implicitly from `__init__()`."""
        super().reset()
        self.within_meta = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:  # noqa: C901
        """Handle opening tags."""
        super().handle_starttag(tag, attrs)
        if not self.within_meta:
            if not self.config[self.M_TAG].fullmatch(tag):
                return
            for attr in attrs:
                if attr[1] is None:
                    return
                if self.config[self.M_ATTR].fullmatch(attr[0]) and self.config[self.M_VALUE].search(attr[1]):
                    logger.debug(Messages.METADATA_MARKER_FOUND.format(attr[1]))
                    self.within_meta = True
                    return
        else:
            if tag == self.K_TAG:
                logger.debug(Messages.METADATA_KEY_MARKER_FOUND.format(tag))
                self.within_k = True
                if self.within_v:
                    # If still processing a value, log about the nesting
                    # error but reset parser so everything starts afresh
                    # like if a new key had been found.
                    logger.debug(Messages.PARSER_NESTING_ERROR_K_IN_V)
                    self.within_v = False
                    self.current_v = self.DEFAULT_V
                return
            if tag == self.V_TAG:
                logger.debug(Messages.METADATA_VALUE_MARKER_FOUND.format(tag))
                self.within_v = True
                if self.within_k:
                    # If a new key is found but a previous one is still
                    # being processed, this nesting error is not fatal
                    # and can still be recovered, up to a certain point.
                    # If some data was got for the key, the parser will
                    # be left in 'within_v' mode to try to get the value
                    # if possible. Otherwise the parser is reset.
                    logger.debug(Messages.PARSER_NESTING_ERROR_V_IN_K)
                    self.within_k = False
                    if not self.current_k:
                        self.within_v = False
                return

    def handle_endtag(self, tag: str) -> None:
        """Handle closing tags."""
        super().handle_endtag(tag)
        if self.within_meta and self.config[self.M_TAG].fullmatch(tag):
            self.within_meta = False
            return
        if self.within_k and tag == self.K_TAG:
            self.within_k = False
            return
        if self.within_v and tag == self.V_TAG:
            self.within_v = False
            self.store_metadata()
            return


class Profile(NamedTuple):
    """Abstraction for profiles."""  # noqa: D204
    url_pattern: re.Pattern[str]
    parser: BaseParser
    parser_config: dict[str, re.Pattern[str]]


def error(message: str, details: str='') -> None:
    """Preprocess and log error *message*, including optional *details*.

    Both an error marker and a header are prepended to *message*, and a
    visual separator is prepended to *details*. In addition to this both
    *message* and *details* are indented.

    Finally, everything is logged using `logger.error()`.
    """
    logger.set_indent(0)
    logger.error(Messages.ERROR_HEADER)

    logger.set_indent(Constants.ERROR_PAYLOAD_INDENT)
    logger.error(message)

    if details := details.strip():
        logger.error(Messages.ERROR_DETAILS_HEADING)
        logger.error('\n'.join(f'{Messages.ERROR_DETAILS_PREAMBLE}{line}' for line in details.split('\n')))
        logger.error(Messages.ERROR_DETAILS_TAIL)
    logger.set_indent(0)


def warning(message: str) -> None:
    """Preprocess and log warning *message*."""
    logger.warning('%s', Messages.WARNING_HEADER + message[0].lower() + message[1:])


def is_accepted_url(value: str | None) -> bool:
    """Check whether value is an accepted URL or not."""
    try:
        # The check is quite crude but it is simple and works perfectly
        # for the program's needs, which are simple, too.
        return urlparse(value).scheme in Constants.ACCEPTED_URL_SCHEMES
    except ValueError:
        return False


def generate_sinkfile_path(base_path: Path) -> Path:
    """Generate a path usable for a data sink, from *base_path*."""
    return base_path.with_stem(base_path.stem + Constants.SINKFILE_STEM)


class WFKStatuses(IntEnum):
    """Return statuses for `wait_for_keypress()`."""  # noqa: D204
    IMPORTED = auto()
    NO_CONSOLE_ATTACHED = auto()
    NO_CONSOLE_TITLE = auto()
    NO_TRANSIENT_FROZEN = auto()
    NO_TRANSIENT_PYTHON = auto()
    WAIT_FOR_KEYPRESS = auto()

def wait_for_keypress() -> WFKStatuses:
    """Wait for a keypress to continue in particular circumstances.

    If `sys.stdout` has an actual console attached **AND** is transient,
    this function will print a simple message telling the end user that
    the program is paused until any key is pressed.
    """
    # First of all, if the script is being imported rather than run then
    # the program must NOT pause. Absolutely NOT.
    if __name__ != '__main__':
        return WFKStatuses.IMPORTED

    # If no console is attached, the program must NOT pause.
    #
    # Since 'sys.stdout.isatty()' returns 'True' under Windows when the
    # 'sys.stdout' stream is redirected to 'NUL', another check, a bit
    # more compicated, is needed here. The test below has been adapted
    # from https://stackoverflow.com/a/33168697
    if not windll.kernel32.GetConsoleMode(get_osfhandle(sys.stdout.fileno()), byref(c_uint())):
        return WFKStatuses.NO_CONSOLE_ATTACHED

    # If there is an attached console, then the program must pause ONLY
    # if that console will automatically close when the program exists.
    # In other words, pause only if the console is transient.
    #
    # Determining if a console is transient is not easy as there is no
    # bulletproof method available for every possible situation.
    #
    # There are TWO main scenarios, though: a frozen executable and a
    # '.py' file. In both cases, the console title has to be obtained.
    buffer_size = MAX_PATH_LEN + 1
    console_title = create_unicode_buffer(buffer_size)
    if not windll.kernel32.GetConsoleTitleW(console_title, buffer_size):
        return WFKStatuses.NO_CONSOLE_TITLE
    console_title = console_title.value

    # If the console is not transient, return, do not pause.
    #
    # For a frozen executable, it is easier: if the console title is not
    # equal to 'sys.executable', then the console is NOT transient.
    #
    # For a '.py' file, this is more complicated, but in most cases if
    # the console title contains the name of the '.py' file, the console
    # is NOT a transient console.
    if getattr(sys, 'frozen', False):
        if console_title != sys.executable:
            return WFKStatuses.NO_TRANSIENT_FROZEN
    elif Constants.APP_NAME in console_title:
        return WFKStatuses.NO_TRANSIENT_PYTHON

    sys.stdout.flush()
    sys.stdout.write(Messages.PRESS_ANY_KEY)
    sys.stdout.flush()
    getch()
    return WFKStatuses.WAIT_FOR_KEYPRESS


def excepthook(exc_type: type[BaseException], exc_value: BaseException, exc_traceback: TracebackType | None) -> None:
    """Log unhandled exceptions. Default exception hook.

    Unhandled exceptions are logged, using the provided arguments, that
    is, the exception type (*exc_type*), its value (*exc_value*) and the
    associated traceback (*exc_traceback*).

    For `OSError` exceptions a different format is used, which includes
    any additional `OSError` information, and no traceback is logged.

    For any other exception, a generic message is logged together with
    the traceback, if available.
    """
    if isinstance(exc_value, OSError):
        message = Messages.UNEXPECTED_OSERROR
        errno_message = Messages.OSERROR_DETAIL_NA
        if exc_value.errno:
            with contextlib.suppress(IndexError):
                errno_message = errno.errorcode[exc_value.errno]
        details = Messages.OSERROR_DETAILS.format(
            exc_type.__name__,
            errno_message,
            exc_value.winerror or Messages.OSERROR_DETAIL_NA,
            exc_value.strerror,
            Messages.OSERROR_DETAIL_NA if exc_value.filename is None else exc_value.filename,
            Messages.OSERROR_DETAIL_NA if exc_value.filename2 is None else exc_value.filename2,
        )
    else:
        message = Messages.UNHANDLED_EXCEPTION
        args = ''
        for arg in exc_value.args:
            args += Messages.EXCEPTION_DETAILS_ARG.format(type(arg).__name__, arg)
        details = Messages.EXCEPTION_DETAILS.format(exc_type.__name__, str(exc_value), args)
    current_frame_source_path = None
    traceback = ''
    for frame in tb.extract_tb(exc_traceback):
        if current_frame_source_path != frame.filename:
            traceback += Messages.TRACEBACK_FRAME_HEADER.format(frame.filename)
            current_frame_source_path = frame.filename
        frame.name = Constants.APP_NAME if frame.name == Messages.TRACEBACK_TOPLEVEL_FRAME else frame.name
        traceback += Messages.TRACEBACK_FRAME_LINE.format(frame.lineno, frame.name, frame.line)
    details += Messages.TRACEBACK_HEADER.format(traceback) if traceback else ''
    error(message, details)


def loggerize(function: Callable[..., ExitCodes]) -> Callable[..., ExitCodes]:
    """Decorate *function* so it runs with logging enabled."""
    @wraps(function)
    def loggerize_wrapper(*args: str) -> ExitCodes:
        logger.config(main_log_output=Constants.MAIN_OUTPUT_PATH, full_log_output=Constants.FULL_OUTPUT_PATH)

        logger.debug(Messages.DEBUGGING_INIT)
        logger.info(Messages.APP_BANNER)
        logger.debug(Constants.USER_AGENT)

        status = function(*args)

        logger.info(Messages.PROCESS_DONE)
        logger.debug(Messages.DEBUGGING_DONE)
        logging.shutdown()
        return status
    return loggerize_wrapper


def keyboard_interrupt_handler(function: Callable[..., ExitCodes]) -> Callable[..., ExitCodes]:
    """Decorate *function* so it handles `KeyboardInterrupt`."""
    @wraps(function)
    def handle_keyboard_interrupt_wrapper(*args: str) -> ExitCodes:
        try:
            return function(*args)
        except KeyboardInterrupt:
            warning(Messages.KEYBOARD_INTERRUPT)
            return ExitCodes.KEYBOARD_INTERRUPT
    return handle_keyboard_interrupt_wrapper


def load_profiles(profiles_path: Path) -> dict[str, Profile]:  # noqa: C901
    """Load the profiles from *profiles_path*.

    Read *profiles_path*, which contains profile definitions, and return
    a processed list of the obtained profiles as a dictionary.

    Dictionary keys are the obtained profile names, while the values are
    the corresponding profile definitions as `Profile` named tuples.

    Raise `MissingProfilesError` if *profiles_path* cannot be opened or
    read.

    Raise `ProfilesSyntaxError` if there are any syntax errors in the
    contents of *profiles_path*.

    The returned dictionary will be empty if no profiles or only empty
    profiles are present in *profiles_path*.
    """
    config = configparser.ConfigParser()
    logger.debug(Messages.LOADING_PROFILES.format(profiles_path))
    try:
        with profiles_path.open(encoding=Constants.UTF8) as inifile:
            config.read_file(inifile)
    except (FileNotFoundError, PermissionError) as exc:
        raise ProfilesError(Messages.MISSING_PROFILES.format(exc.filename)) from exc
    except configparser.Error as exc:
        errorname = type(exc).__name__.removesuffix(configparser.Error.__name__)
        raise ProfilesError(Messages.PROFILES_WRONG_SYNTAX.format(errorname), exc) from exc

    profiles: dict[str, Profile] = {}
    parsers = [parser() for parser in BaseParser.__subclasses__()]
    for section in config.sections():
        if not config[section]:
            continue
        parser_config: dict[str, re.Pattern[str]] = {}
        for key, value in config[section].items():
            if not value:
                continue
            try:
                parser_config[key] = re.compile(value, re.IGNORECASE)
            except re.error as exc:
                message = Messages.PROFILES_WRONG_SYNTAX.format(Constants.PROFILE_BAD_REGEX_ERROR)
                details = Messages.PROFILES_WRONG_SYNTAX_DETAILS.format(
                    section, exc.msg,
                    key, Messages.PROFILES_WRONG_SYNTAX_DETAILS_SEPARATOR, exc.pattern,
                    '', (exc.pos or 0) + len(key) + len(Messages.PROFILES_WRONG_SYNTAX_DETAILS_SEPARATOR),
                    # The empty string above is needed as a placeholder.
                )
                raise ProfilesError(message, details) from exc
        url_pattern = parser_config.pop(Constants.PROFILE_URL_PATTERN_KEY, None)
        if url_pattern is None:
            raise ProfilesError(Messages.INVALID_PROFILE.format(section), Messages.PROFILE_WITHOUT_URL)
        for parser in parsers:
            if parser_config.keys() == parser.PARAMETERS:
                break
        else:
            raise ProfilesError(Messages.INVALID_PROFILE.format(section))
        profiles[section] = Profile(url_pattern, parser, parser_config)
    if not profiles:
        raise ProfilesError(Messages.EMPTY_PROFILES.format(profiles_path))
    return profiles


def parse_arguments(*args: str) -> Generator[tuple[str, Handler]]:
    """Parse **args*.

    Parse each argument in **args*, check if it is a valid source, then
    identify its type and build the corresponding handler.

    Yield tuple containing the source and its corresponding handler.
    """
    for arg in args:
        logger.debug(Messages.PROCESSING_ARG.format(arg))
        if is_accepted_url(arg):
            logger.debug(Messages.ARG_IS_SOURCE_SINGLE_URL)
            handler = single_url_handler(arg)
        elif arg.endswith(Constants.TEXTFILE_SUFFIX):
            logger.debug(Messages.ARG_IS_SOURCE_TEXTFILE)
            handler = textfile_handler(Path(arg))
        elif arg.endswith(Constants.SPREADSHEET_SUFFIX):
            logger.debug(Messages.ARG_IS_SOURCE_SPREADSHEET)
            handler = spreadsheet_handler(Path(arg))
        else:
            logger.debug(Messages.ARG_IS_SOURCE_UNSUPPORTED)
            handler = unsupported_source_handler()
        yield arg, handler


# NOTE: the handlers below have to be generators for two reasons.
#
# The first reason, the list of URLs retrieved by the handler from the
# source can be potentially large. So, using a generator is a lot more
# efficient.
#
# The second reason, a generator allows the caller to use the handler in
# two separate stages: URL is retrieved first, so the handler can keep
# processing it, and then the obtained metadata for the URL is obtained
# and sent back to the handler. In short, the handler is acting as both
# a coroutine and a generator.
#
# Since the metadata is sent back to the handler, it triggers another
# iteration so a double yield is needed. The first one is the one that
# yields the URLs and the second one yields a response to the caller
# after the metadata has been received back into the handler.
#
# The first 'yield Constants.HANDLER_BOOTSTRAP_SUCCESS' expression in
# the handlers is for signalling successful initialization after priming
# the generator/coroutine.

def unsupported_source_handler() -> Handler:
    """Handle unsupported sources."""
    raise SourceError(Messages.UNSUPPORTED_SOURCE)
    # The 'yield' below is needed, otherwise the return value of the
    # handler will not be a generator/coroutine and that is needed so
    # this handler is compatible with all the others. Of course, since
    # the handler just raises an exception, the yield will never be
    # executed…
    yield Constants.HANDLER_BOOTSTRAP_SUCCESS  # pylint: disable=unreachable


def single_url_handler(url: str) -> Handler:
    """Handle single *url*.

    The metadata for *url* is logged with `logging.INFO` level, so it
    will be printed on `sys.stdout` in addition to the corresponding log
    file, and will be also written as key-value pairs into a dump text
    file whose name is based on a properly sanitized version of *url*.

    The dump output file has UTF-8 encoding.
    """
    sinkfile_path = generate_sinkfile_path(url_to_path(url).with_suffix(Constants.TEXTFILE_SUFFIX))
    with sinkfile_path.open('w', encoding=Constants.UTF8) as sink:
        logger.debug(Messages.DUMPING_METADATA_TO_SINK.format(sinkfile_path))
        yield Constants.HANDLER_BOOTSTRAP_SUCCESS
        if is_accepted_url(url):
            metadata = yield url
            yield url
            if metadata:
                sink.write(Constants.TEXTSINK_METADATA_HEADER.format(url))
                for key, value in metadata.items():
                    logger.debug(Messages.DUMPING_METADATA_K_V.format(key, value))
                    message = Constants.TEXTSINK_METADATA_PAIR.format(key, value)

                    # Console output allowed here because it is part of
                    # the handler documented and expected behavior.
                    logger.indent()
                    logger.info(message)
                    logger.dedent()

                    sink.write(message)
                sink.write(Constants.TEXTSINK_METADATA_FOOTER)


def url_to_path(url: str) -> Path:
    """Convert the given *url* to a valid path.

    The method is quite crude but it works: replace all ASCII non-word
    characters (potentially unsafe if they are present in a path) by a
    visually unobtrusive character which is safe to use in a path, so
    the path is still readable.
    """
    return Path(re.sub(Constants.URL_UNSAFE_CHARS_RE, Constants.URL_UNSAFE_REPLACE_CHAR, url, flags=re.ASCII))


def textfile_handler(source_file: Path) -> Handler:
    """Handle URLs from *source_file* text file, one per line.

    The metadata for each URL is dumped into another text file, named
    after *source_file*. In that output file, first the URL is written,
    then the metadata as key-value pairs.

    All files are assumed to have UTF-8 encoding.
    """
    sinkfile_path = generate_sinkfile_path(source_file)
    encoding = Constants.UTF8
    with source_file.open(encoding=encoding) as source, sinkfile_path.open('w', encoding=encoding) as sink:
        logger.debug(Messages.DUMPING_METADATA_TO_SINK.format(sinkfile_path))
        yield Constants.HANDLER_BOOTSTRAP_SUCCESS
        for line in source.readlines():
            url = line.strip()
            if not is_accepted_url(url):
                continue
            metadata = yield url
            yield url
            if metadata:
                sink.write(Constants.TEXTSINK_METADATA_HEADER.format(url))
                for key, value in metadata.items():
                    logger.debug(Messages.DUMPING_METADATA_K_V.format(key, value))
                    sink.write(Constants.TEXTSINK_METADATA_PAIR.format(key, value))
                sink.write(Constants.TEXTSINK_METADATA_FOOTER)


def spreadsheet_handler(source_file: Path) -> Handler:
    """Handle URLs from spreadsheet *source_file*, one per row. Ish.

    The metadata obtained is dumped into another spreadsheet, whose name
    is based upon *source_file*. It is not created anew, it is a copy of
    the *source_file* spreadsheet and later modified.

    The metadata is added to the destination spreadsheet in new columns,
    one per metadata. These columns are marked with a prefix so they are
    easier to recognize.

    NOTE: not all sheets are processed, only the first one because it is
    the one where the URLs for the items are. Allegedly…s
    """
    sinkfile_path = generate_sinkfile_path(source_file)
    logger.debug(Messages.COPYING_WORKBOOK.format(sinkfile_path))

    copy2(source_file, sinkfile_path)
    try:
        source_workbook = load_workbook(source_file)
    except (KeyError, BadZipFile) as exc:
        details = str(exc).strip(Constants.DOUBLE_QUOTE_CHAR)
        details = details[0].lower() + details[1:]
        raise SourceError(Messages.SOURCE_SHEET_IS_INVALID, details) from exc
    sink_workbook = load_workbook(sinkfile_path)
    yield Constants.HANDLER_BOOTSTRAP_SUCCESS

    source_sheet = source_workbook.worksheets[0]
    logger.debug(Messages.WORKING_SHEET.format(source_sheet.title))

    sink_sheet = sink_workbook.worksheets[0]

    logger.debug(Messages.INSERTING_HEADING_ROW)
    sink_sheet.insert_rows(1, 1)

    for row in source_sheet.rows:
        row_number = row[0].row
        logger.debug(Messages.PROCESSING_ROW.format(row_number))
        if (url := get_url_from_row(row)) is None:
            continue
        metadata = yield url
        yield url
        if metadata and row_number:
            store_metadata_in_sheet(sink_sheet, row_number, metadata)
    sink_workbook.save(sinkfile_path)
    sink_workbook.close()
    source_workbook.close()


def get_url_from_row(row: tuple[Cell | MergedCell, ...]) -> str | None:
    """Find first URL in row."""
    url = None
    for cell in row:
        if cell.data_type != CELLTYPE_STRING:
            logger.debug(Messages.NONSTRING_CELL.format(cell.coordinate))
            continue
        if is_accepted_url(str(cell.value)):
            url = str(cell.value)
            logger.debug(Messages.URL_FOUND_IN_CELL.format(cell.coordinate, url))
            break  # Only the FIRST URL found in each row is considered.
    return url


def store_metadata_in_sheet(
    sheet: Worksheet,
    row: int,
    metadata: dict[str, str],
    static: SimpleNamespace = SimpleNamespace(known_metadata = {}),  # noqa: B008
) -> None:
    """Store *metadata* in provided *sheet* at given *row*.

    For new metadata, a new column is added to the *sheet*. For already
    existing metadata, the value is added to the existing column.
    """
    # NOTE: since default parameters are evaluated just once, a typical
    # trick for simulating static variables in functions is using a fake
    # default parameter and using 'SimpleNamespace':
    # https://stackoverflow.com/a/51437838
    if not metadata:
        return
    for key, value in metadata.items():
        if key not in static.known_metadata:
            column_header = Constants.SPREADSHEET_METADATA_COLUMN_TITLE.format(key)
            logger.debug(Messages.NEW_METADATA_FOUND.format(key))
            column = sheet.max_column + 1
            static.known_metadata[key] = column
            logger.debug(Messages.METADATA_STORED_IN_COLUMN.format(key, get_column_letter(column)))
            cell = sheet.cell(row=1, column=column, value=column_header)
            cell.font = Font(name=Constants.SPREADSHEET_CELL_FONT)
            cell.fill = PatternFill(fgColor=Constants.SPREADSHEET_CELL_COLOR, fill_type=Constants.SPREADSHEET_CELL_FILL)
            # Set column width.
            #
            # As per Excel specification, the width units are the width
            # of the zero character of the font used by the Normal style
            # for a workbook. So a column of width 10 would fit exactly
            # 10 zero characters in the font specified by the 'Normal'
            # style.
            #
            # No, no kidding.
            #
            # Since this width units are, IMHO, totally arbitrary, let's
            # choose an arbitrary column width, so 42.
            #
            # To wit, the Answer to the Ultimate Question of Life, the
            # Universe, and Everything.
            sheet.column_dimensions[get_column_letter(column)].width = 42
            # This is needed because in some occasions Excel files are
            # not properly generated and the last column has a 'max'
            # field too large, and that has an unintended consequence:
            # ANY change to the settings of that column affects ALL the
            # following ones whose index is less than 'max'… So, it's
            # better to fix that field.
            sheet.column_dimensions[get_column_letter(column)].max = column
        logger.debug(Messages.DUMPING_METADATA_K_V.format(key, value))
        # Since a heading row is inserted, the rows where metadata has
        # to go have now an +1 offset, as they have been displaced.
        sheet.cell(row + 1, static.known_metadata[key], value=value)


def bootstrap(handler: Handler) -> None:
    """Bootstrap *handler* and deal with initialization errors."""
    try:
        handler.send(None)
    except FileNotFoundError as exc:
        raise SourceError(Messages.INPUT_FILE_NOT_FOUND) from exc
    except PermissionError as exc:
        if Path(exc.filename).stem.endswith(Constants.SINKFILE_STEM):
            raise SourceError(Messages.OUTPUT_FILE_NO_PERMISSION) from exc
        raise SourceError(Messages.INPUT_FILE_NO_PERMISSION) from exc


def get_parser(url: str, profiles: dict[str, Profile]) -> BaseParser:
    """Return the appropriate parser for the *url*.

    The appropriate parser is retrieved by finding the needed profile
    within *profiles* whose 'url_pattern' matches the given *url*. Then,
    the parser is returned after being properly configured to handle the
    *url* contents.
    """
    for profile_name, profile in profiles.items():
        if profile.url_pattern.search(url):
            logger.debug(Messages.DETECTED_PROFILE.format(profile_name))
            profile.parser.configure(profile.parser_config)
            return profile.parser
    raise SkimmingError(Messages.NO_MATCHING_PROFILE)


def saca_las_mantecas(url: str, parser: BaseParser) -> dict[str, str]:  # noqa: C901
    """Saca las mantecas.

    Saca las mantecas from the given *url*, that is, retrieve contents,
    parse them using *parser*, and obtain library catalogue metadata, if
    any.

    Return obtained metadata as a dictionary.
    """
    try:
        contents, encoding = retrieve_url(url)
    except URLError as exc:
        # Depending on the particular error which happened, the reason
        # attribute of the 'URLError' exception can be a simple error
        # message, an instance of some 'OSError' derived Exception, etc.
        # So, in order to have useful messages, some discrimination has
        # to be done here.
        if isinstance(exc.reason, OSError):
            error_code = Messages.UNKNOWN_ERRNO
            if exc.reason.errno:
                try:
                    error_code = errno.errorcode[exc.reason.errno]
                except KeyError:
                    error_code = exc.reason.errno
            error_reason = exc.reason.strerror or ''
            details = Messages.OSLIKE_URLERROR
        elif isinstance(exc, HTTPError):
            error_code = exc.code
            error_reason = exc.reason.lower()
            details = Messages.HTTP_PROTOCOL_URLERROR
        else:
            error_code = ''
            error_reason = str(exc.reason)
            details = Messages.GENERIC_URLERROR
        error_reason = (error_reason[0].lower() + error_reason[1:]).rstrip(Constants.PERIOD)
        raise SkimmingError(Messages.URL_ACCESS_ERROR, details.format(error_code, error_reason)) from exc
    # Apparently, 'HTTPException', 'ConnectionError' and the subclasses
    # are masked or wrapped by 'urllib', and documentation is not very
    # informative. So, just in case something weird happen, it is better
    # to handle these exception types here as well.
    except HTTPException as exc:
        details = f'{type(exc).__name__}: {exc}.'
        raise SkimmingError(Messages.HTTP_RETRIEVAL_ERROR, details) from exc
    except ConnectionError as exc:
        error_code = Messages.UNKNOWN_ERRNO
        if exc.errno:
            with contextlib.suppress(IndexError):
                error_code = errno.errorcode[exc.errno]
        details = ''
        if exc.strerror:
            details = f'{exc.strerror.capitalize().rstrip(".")}.'
        raise SkimmingError(Messages.CONNECTION_ERROR.format(error_code), details) from exc

    if not contents:
        raise SkimmingError(Messages.NO_CONTENTS_ERROR)

    parser.feed(contents.decode(encoding))
    parser.close()
    if metadata := parser.get_metadata():
        return metadata
    raise SkimmingError(Messages.NO_METADATA_FOUND)


def retrieve_url(url: str) -> tuple[bytes, str]:
    """Retrieve contents from *url*.

    First resolve any `meta http-equiv="Refresh"` redirection for *url*
    and then get the contents as a byte string.

    Then, detect the contents encoding (in HTML jargon, the charset).

    Return a (contents, charset) tuple.
    """
    if not is_accepted_url(url):
        raise URLError(Messages.UNKNOWN_URL_TYPE.format(url))

    current_url: str | None = url

    if url.startswith(Constants.FILE_SCHEME):
        current_url = resolve_file_url(url)

    contents = b''
    charset = ''
    headers = {Constants.USER_AGENT_HEADER: Constants.USER_AGENT}
    while current_url:
        logger.debug(Messages.PROCESSING_URL.format(current_url))
        with urlopen(Request(current_url, headers=headers)) as response:  # noqa: S310
            # First, check if any redirection is needed and try to get
            # the charset the easy way.
            contents = response.read()
            charset = response.headers.get_content_charset()
        current_url = get_redirected_url(contents, current_url)

    # In this point, we have the contents as a byte string.
    # If the charset is 'None', it has to be determined the hard way.
    if not charset:
        logger.debug(Messages.CHARSET_NOT_IN_HEADERS)
        charset = detect_html_charset(contents)
    else:
        logger.debug(Messages.CHARSET_IN_HEADERS)
    logger.debug(Messages.CONTENTS_ENCODING.format(charset))

    return contents, charset


def resolve_file_url(url: str) -> str:
    """Resolve relative paths in `file:` *url*."""
    parsed_url = urlparse(url)
    resolved_path = unquote(parsed_url.path[1:])
    resolved_path = Path(resolved_path).resolve().as_posix()
    resolved_path = quote(resolved_path, safe=Constants.FILE_URL_SAFE_CHARS)
    return parsed_url._replace(path=Constants.FILE_URL_SEPARATOR + resolved_path).geturl()


def get_redirected_url(contents: bytes, base_url: str) -> str | None:
    """Get redirected URL, if any, from *contents*.

    Get redirected URL from a `meta http-equiv="refresh"` pragma in the
    contents. Use *base_url* as base URL for redirection, if some parts
    are missing in the URL specified by the pragma.

    Return redirected URL, or `None` if there is no redirection pragma.
    """
    if match := re.search(Constants.META_REFRESH_RE, contents, re.IGNORECASE):
        parsed_url = urlparse(base_url)
        redirected_url = urlparse(match.group(1).decode(Constants.ASCII))
        for field in parsed_url._fields:
            value = getattr(parsed_url, field)
            # If not specified in the obtained redirected URL, both the
            # scheme and netloc will be reused from the base URL. Any
            # other field will be obtained from the redirected URL and
            # used, no matter if it is empty.
            if field in Constants.URL_FIELDS_TO_REUSE and not getattr(redirected_url, field):
                redirected_url = redirected_url._replace(**{field: value})
        redirected_url = urlunparse(redirected_url)
        logger.debug(Messages.REDIRECTED_URL.format(redirected_url))
        return redirected_url
    return None


def detect_html_charset(contents: bytes) -> str:
    """Obtain charset for *content* from HTML tags, if any.

    If the charset can not be determined a sane fallback is used. It may
    look like UTF-8 would be such a sane fallback, because some webpages
    may NOT specify any encoding if they are using UTF-8 and it happens
    to be identical to ASCII for 7-bit codepoints, **BUT** the problem
    is that UTF-8 will fail for webpages whose encoding is any ISO/IEC
    8859 variant.

    So, the sane default is another, set in the global configuration,
    and it is based on the encoding most frequently used by the webpages
    this program will generally process.
    """
    charset = Constants.FALLBACK_HTML_CHARSET
    if match := re.search(Constants.META_HTTP_EQUIV_CHARSET_RE, contents, re.IGNORECASE):
        # Next best thing, from the meta http-equiv="content-type".
        logger.debug(Messages.CHARSET_FROM_HTTP_EQUIV)
        charset = match.group(1).decode(Constants.ASCII)
    elif match := re.search(Constants.META_CHARSET_RE, contents, re.IGNORECASE):
        # Last resort, from some meta charset, if any…
        logger.debug(Messages.CHARSET_FROM_META_CHARSET)
        charset = match.group(1).decode(Constants.ASCII)
    else:
        logger.debug(Messages.CHARSET_FROM_DEFAULT)
    return charset


@loggerize
@keyboard_interrupt_handler
def main(*args: str) -> ExitCodes:
    """."""
    exitcode = ExitCodes.SUCCESS

    if not args:
        # Input arguments should be provided automatically when program
        # is used as a drag'n'drop target which is actually the intended
        # way of operation, generally speaking.
        #
        # But the program can be also run by hand from a command prompt,
        # so it is better to signal the end user with an error and give
        # an explanation if no input arguments are provided, as soon as
        # possible.
        error(Messages.NO_ARGUMENTS)
        return ExitCodes.NO_ARGUMENTS

    try:
        profiles = load_profiles(Constants.INIFILE_PATH)
        logger.debug(Messages.FOUND_PROFILES.format(Constants.OUTPUT_SEPARATOR.join(profiles.keys())))
    except ProfilesError as exc:
        error(str(exc), str(exc.details) if exc.details else '')
        return ExitCodes.ERROR

    logger.info(Messages.SKIMMING_MARKER)
    logger.indent()
    for source, handler in parse_arguments(*args):
        logger.info(Messages.SOURCE_LABEL.format(source))
        try:
            bootstrap(handler)
        except SourceError as exc:
            logger.indent()
            warning(str(exc))
            logger.dedent()
            exitcode = ExitCodes.WARNING
            continue

        logger.indent()
        for url in handler:
            logger.info(url)
            metadata = {}
            try:
                parser = get_parser(url, profiles)
                metadata = saca_las_mantecas(url, parser)
            except SkimmingError as exc:
                logger.indent()
                warning(str(exc))
                logger.debug(exc.details)
                logger.dedent()
                exitcode = ExitCodes.WARNING
            finally:
                # No matter if metadata has actual contents or not, the
                # handler has to be 'advanced' to the next URL, so the
                # metadata it is expecting has to be sent back to the
                # handler.
                handler.send(metadata)
        logger.dedent()
    logger.dedent()
    return exitcode


if __name__ == '__main__':
    atexit.register(wait_for_keypress)
    sys.excepthook = excepthook
    sys.exit(main(*sys.argv[1:]))
