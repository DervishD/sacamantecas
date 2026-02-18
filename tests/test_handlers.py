#! /usr/bin/env python3
"""Test suite for the different handlers of sources and sinks."""
from hashlib import algorithms_available, new as new_hash
from pathlib import Path
from random import choice, randrange
from typing import TYPE_CHECKING
from uuid import uuid4

from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import pytest

from sacamantecas import (
    bootstrap,
    Constants,
    get_url_from_row,
    Handler,
    Messages,
    single_url_handler,
    SourceError,
    spreadsheet_handler,
    textfile_handler,
    url_to_path,
)

if TYPE_CHECKING:
    from collections.abc import Callable

HASHES = [hash_function for hash_function in algorithms_available if not hash_function.startswith('shake')]
SAMPLE_URLS = [f'{choice(Constants.ACCEPTED_URL_SCHEMES)}://subdomain{i}.domain.tld' for i in range(10)]  # noqa: S311
EXPECTED_METADATA = {u: {h: new_hash(h, u.encode(Constants.UTF8)).hexdigest() for h in HASHES} for u in SAMPLE_URLS}


def test_single_url_handler(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:  # pylint: disable=unused-variable
    """Test single URLs."""
    single_url = url_to_path('url://subdomain.domain.toplevel/path?param1=value1&param2=value2')
    expected = Path('url___subdomain_domain_toplevel_path_param1_value1_param2_value2')

    assert single_url == expected

    sinkfile_path = tmp_path / f'testsink{Constants.SINKFILE_STEM}.txt'
    def patched_generate_sinkfile_path(_: Path) -> Path:
        return sinkfile_path

    monkeypatch.setitem(single_url_handler.__globals__, 'generate_sinkfile_path', patched_generate_sinkfile_path)

    handler = single_url_handler(SAMPLE_URLS[0])
    bootstrap(handler)

    urls: list[str] = []
    for url in handler:
        assert url is not None
        assert not isinstance(url, bool)

        handler.send(EXPECTED_METADATA[url])

        urls.append(url)


    assert sinkfile_path.is_file()
    assert len(urls) == 1
    assert urls[0] == SAMPLE_URLS[0]

    result = sinkfile_path.read_text().rstrip(Constants.TEXTSINK_METADATA_FOOTER).splitlines()

    assert result[0] == SAMPLE_URLS[0]

    result = dict(line.strip().split(Constants.TEXTSINK_METADATA_SEPARATOR) for line in result[1:])

    assert result == EXPECTED_METADATA[urls[0]]


def test_textfile_handler(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:  # pylint: disable=unused-variable
    """Test textfile handler."""
    sourcefile_path = tmp_path / 'urls.txt'
    sourcefile_path.write_text('\n'.join(SAMPLE_URLS), encoding='utf-8')
    sinkfile_path = tmp_path / f'testsink{Constants.SINKFILE_STEM}.txt'
    def patched_generate_sinkfile_path(_: Path) -> Path:
        return sinkfile_path

    monkeypatch.setitem(textfile_handler.__globals__, 'generate_sinkfile_path', patched_generate_sinkfile_path)

    handler = textfile_handler(sourcefile_path)
    bootstrap(handler)

    urls: list[str] = []
    for url in handler:
        assert url is not None
        assert not isinstance(url, bool)

        handler.send(EXPECTED_METADATA[url])

        urls.append(url)

    assert sinkfile_path.is_file()
    assert len(urls) == len(SAMPLE_URLS)
    assert urls == SAMPLE_URLS

    result: dict[str, dict[str, str]] = {}
    current_k = None
    for line in sinkfile_path.read_text().rstrip(Constants.TEXTSINK_METADATA_FOOTER).splitlines():
        if not line.rstrip():
            continue
        if line.startswith(Constants.TEXTSINK_METADATA_INDENT) and current_k is not None:
            result[current_k].update(dict([line.strip().split(Constants.TEXTSINK_METADATA_SEPARATOR)]))
            continue
        current_k = line.strip()
        result[current_k] = {}

    assert result == EXPECTED_METADATA


FAKE_METADATA_COLUMNS = 10
# pylint: disable-next=unused-variable,too-many-locals
def test_spreadsheet_handler(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """Test spreadsheet handler."""
    sourcefile_path = tmp_path / 'urls.xlsx'
    headings = [f'Heading_{i}' for i in range(FAKE_METADATA_COLUMNS)]

    workbook = Workbook()
    sheet = workbook.active

    assert sheet is not None
    assert isinstance(sheet, Worksheet)

    sheet.append(headings)

    for column in range(FAKE_METADATA_COLUMNS):
        sheet.column_dimensions[get_column_letter(column + 1)].width = 33

    for url in SAMPLE_URLS:
        row = [uuid4().hex[:randrange(5, 20)] for _ in range(FAKE_METADATA_COLUMNS)]  # noqa: S311
        row.insert(randrange(len(row)), url)  # noqa: S311
        sheet.append(row)

    workbook.save(sourcefile_path)
    workbook.close()

    sinkfile_path = tmp_path / f'testsink{Constants.SINKFILE_STEM}.xlsx'
    def patched_generate_sinkfile_path(_: Path) -> Path:
        return sinkfile_path

    monkeypatch.setitem(spreadsheet_handler.__globals__, 'generate_sinkfile_path', patched_generate_sinkfile_path)

    handler = spreadsheet_handler(sourcefile_path)
    bootstrap(handler)

    urls: list[str] = []
    for url in handler:
        assert isinstance(url, str)

        handler.send(EXPECTED_METADATA[url])

        urls.append(url)

    assert sinkfile_path.is_file()
    assert len(urls) == len(SAMPLE_URLS)
    assert urls == SAMPLE_URLS

    result = {}
    workbook = load_workbook(sinkfile_path)
    sheet = workbook.worksheets[0]

    assert isinstance(sheet, Worksheet)

    headers: dict[int ,str] = {}
    for cell in next(sheet.rows):
        value = str(cell.value)
        if not value or not value.startswith(Constants.SPREADSHEET_METADATA_COLUMN_MARKER):
            continue
        value = value.removeprefix(Constants.SPREADSHEET_METADATA_COLUMN_MARKER)
        assert cell.column is not None
        headers[cell.column] = value

    for row in sheet.rows:
        if (url := get_url_from_row(row)) is None:
            continue
        assert row[0].row is not None
        result[url] = {k: sheet.cell(row[0].row, column).value for column, k in headers.items()}
    workbook.close()

    assert result == EXPECTED_METADATA


@pytest.mark.parametrize(('suffix', 'handler_factory'), [
    ('.txt', textfile_handler),
    ('.xlsx', spreadsheet_handler),
], ids=[
    'test_missing_txt_source',
    'test_missing_xlsx_source',
])
# pylint: disable-next=unused-variable
def test_missing_source(tmp_path: Path, suffix: str, handler_factory: Callable[[Path], Handler]) -> None:
    """Test handling of missing sources."""
    handler = handler_factory(tmp_path / f'non_existent{suffix}')

    with pytest.raises(SourceError) as excinfo:
        bootstrap(handler)

    assert str(excinfo.value).startswith(Messages.INPUT_FILE_NOT_FOUND)


@pytest.mark.parametrize(('unreadable_path', 'handler_factory'), [
    ('unreadable_textfile.txt', textfile_handler),
    ('unreadable_spreadsheet.xlsx', spreadsheet_handler),
], indirect=['unreadable_path'], ids=[
    'test_txt_input_no_permission',
    'test_xlsx_input_no_permission',
])
# pylint: disable-next=unused-variable
def test_input_no_permission(unreadable_path: Path, handler_factory: Callable[[Path], Handler]) -> None:
    """Test handling of unreadable files."""
    handler = handler_factory(unreadable_path)

    with pytest.raises(SourceError) as excinfo:
        bootstrap(handler)

    assert str(excinfo.value).startswith(Messages.INPUT_FILE_NO_PERMISSION)


@pytest.mark.parametrize(('source_stem', 'unwritable_path', 'handler_factory'), [
    ('http://s.url', f'unwritable_single_url{Constants.SINKFILE_STEM}.txt', single_url_handler),
    ('s.txt', f'unwritable_textfile{Constants.SINKFILE_STEM}.txt', textfile_handler),
    ('s.xlsx', f'unwritable_spreadsheet{Constants.SINKFILE_STEM}.xlsx', spreadsheet_handler),
], indirect=['unwritable_path'], ids=[
    'test_url_output_no_permission',
    'test_txt_output_no_permission',
    'test_xlsx_output_no_permission',
])
# pylint: disable-next=unused-variable
def test_output_no_permission(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    source_stem: str,
    unwritable_path: Path,
    handler_factory: Callable[[str | Path], Handler],
) -> None:
    """Test handling of non-writable files."""
    def patched_generate_sinkfile_path(_: Path) -> Path:
        return unwritable_path

    monkeypatch.setitem(handler_factory.__globals__, 'generate_sinkfile_path', patched_generate_sinkfile_path)

    if not source_stem.startswith('http://'):
        source = tmp_path / source_stem
        source.write_text('')
    else:
        source = source_stem

    handler = handler_factory(source)

    with pytest.raises(SourceError) as excinfo:
        bootstrap(handler)

    assert str(excinfo.value).startswith(Messages.OUTPUT_FILE_NO_PERMISSION)
