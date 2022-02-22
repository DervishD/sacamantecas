#! /usr/bin/env python3
"""
Obtains a list of book URIs from an Excel (xls/xlsx) input file, performs the
necessary web redirections, if needed, and then gets some metadata for those
books from the Biblioteca Histórica Municipal of Madrid book catalogue.

In short, the program saca las mantecas...
"""

# Revision 2018-12-13T12:11:57+0100

# Imports
import sys
import os.path
import errno
import logging
import traceback
from shutil import copy2
from urllib.request import urlopen
from urllib.parse import urlparse, urlunparse
from urllib.error import URLError
import re
from html.parser import HTMLParser
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import SheetTitleException, InvalidFileException
from openpyxl.utils.cell import get_column_letter
import win32con
import win32ui

# Pattern for finding the URIs on the input file.
CONFIG_URI_PATTERN = r'https?://catalogos.munimadrid.es'
# Class for the <div> containing keys for the new metadata.
CONFIG_K_DIV_CLASS = 'auth'
# Class for the <div> containing values for the new metadata.
CONFIG_V_DIV_CLASS = 'titn'
# Prefix to add to headers for the columns where the new metadata will go.
CONFIG_PREFIX = '[sm] '
# Heading for the processed items marker column.
CONFIG_MARKER_HEADING = CONFIG_PREFIX + 'Mantecas'
# Marker for the processed items.
CONFIG_MARKER = '✓'

# sys.modules[__name__].__file__ is used to determine the program's fully
# qualified directory and filename, so if it's not defined for some reason
# (which may happen...) it's better to break execution here.
try:
    PROGRAM = os.path.realpath(sys.modules[__name__].__file__)
    PROGRAM_NAME = os.path.splitext(os.path.basename(PROGRAM))[0]
    PROGRAM_PATH = os.path.dirname(PROGRAM)
except NameError:
    sys.exit('Error de inicialización del programa.')


# Helper funcion just in case another way of signalling the end user is needed,
# instead of a simple, ugly but nonetheless useful Windows MessageBox.
def error(message):
    """
    Shows the error message 'message' on a Windows API MessageBox.
    """
    win32ui.MessageBox(message, f'Error en {PROGRAM_NAME}', win32con.MB_ICONERROR)


# Install the exception hook.
sys.excepthook = lambda exc_type, exc_value, exc_traceback: (
    error(f'Excepción sin gestionar: "{exc_type.__name__}"\n' +
          'at file "{}", line {}\n\n'.format(*traceback.extract_tb(exc_traceback)[-1]) +
          ''.join(traceback.format_exception(exc_type, exc_value, exc_traceback)))
)


class LibraryCatalogueExcel():
    """
    Represents an Excel file containing some information about items belonging
    to library catalogues.
    Originally created for the Biblioteca Histórica Municipal of Madrid.
    """
    def __init__(self, filename):
        """
        Loads the 'filename' Excel workbook.
        """
        self.filename = filename
        self.workbook = load_workbook(filename)
        # Keys are metadata names, values are the column where that metadata is stored.
        self.metadata_columns = {}
        # Style for cells on the header row
        self.heading_style = {
            'font': Font(name='Calibri'),
            'fill': PatternFill(start_color='baddad', fill_type='solid'),
        }
        # NOTE: not all sheets are processed, only the first one because it is
        # (allegedly) the one where the URIs for the items are.
        self.sheet = self.workbook.worksheets[0]
        logging.debug('La hoja con la que se trabajará es «%s»".', self.sheet.title)

        # Add the already present metadata and columns
        header_row = [cell.value for cell in list(self.sheet.iter_rows(min_row=1, max_row=1))[0]]
        try:
            start = header_row.index(CONFIG_MARKER_HEADING)
            logging.debug('Se encontró una columna de marcadores: «%s».', get_column_letter(start))
            for i, metadata in enumerate(header_row[start:], start=start+1):
                self.metadata_columns[metadata] = i
        except ValueError:
            self.__add_metadata_heading(CONFIG_MARKER_HEADING)
            marker_column = get_column_letter(self.metadata_columns[CONFIG_MARKER_HEADING])
            self.sheet.column_dimensions[marker_column].hidden = True
            logging.debug('Añadiendo columna oculta de marcadores: «%s».', marker_column)

    def find(self, pattern):
        """
        Generator. Finds on the default worksheet cells whose value matches
        regex 'pattern', returning only the first matching cell for each row.
        There's no need to match all cells in each row, as the first one
        containing and URI is the correct one.
        """
        for row in self.sheet.rows:
            for cell in row:
                logging.debug('%s [%s] «%s».', cell, cell.data_type, cell.value)
                if cell.data_type != 's':
                    continue
                if re.match(pattern, cell.value):
                    yield cell
                    break  # Return first matching cell ONLY.

    def __add_metadata_heading(self, heading):
        """
        Adds the new key to first row of sheet, as a heading
        """
        column = self.sheet.max_column + 1
        self.metadata_columns[heading] = column
        logging.debug('El metadato «%s» irá en la columna «%s».', heading, get_column_letter(column))
        cell = self.sheet.cell(row=1, column=column, value=heading)
        cell.font = self.heading_style['font']
        cell.fill = self.heading_style['fill']
        # Set column width. As per Excel specification, the width units are
        # the width of the zero character of the font used by the Normal
        # style for a workbook. So a column of width 10 would fit exactly
        # 10 zero characters in the font specified by the Normal style.
        # No, no kidding.
        # Since this width units are, IMHO, totally arbitrary, let's choose
        # an arbitrary column width. To wit, the Answer to the Ultimate
        # Question of Life, the Universe, and Everything.
        self.sheet.column_dimensions[get_column_letter(column)].width = 42
        # This is needed because sometimes Excel files are not properly
        # generated and the last column has a 'max' field too large, and that
        # has an unintended consequence: any change to the settings of that
        # column affects ALL the following ones whose index is less than
        # 'max'... Just in case that happens, and since some columsn are going
        # to be hidden, it's better to fix that field.
        self.sheet.column_dimensions[get_column_letter(column)].max = column

    def add_metadata_for_cell(self, cell, metadata):
        """
        Adds the metadata specified by 'key' and 'value' to the default
        worksheet at the specified row, in a new column if the metadata key
        doesn't already exists on the sheet.

        Adds the header, also, if it doesn't exist, and styles the header.
        NOTE: the styling is just a best effort, and it's fragile. It depends
        on the sheet having headers on the first row, and the style used is
        that of the FIRST header.
        """
        for key, value in metadata.items():
            key = CONFIG_PREFIX + key
            if key not in self.metadata_columns:
                logging.debug('Se encontró un metadato nuevo, «%s».', key)
                self.__add_metadata_heading(key)
            # Add the value to the proper column
            logging.debug('Añadiendo metadato «%s» con valor «%s».', key, value)
            self.sheet.cell(cell.row, self.metadata_columns[key], value=value)
        self.sheet.cell(cell.row, self.metadata_columns[CONFIG_MARKER_HEADING], value=CONFIG_MARKER)

    def is_processed(self, cell):
        """
        Returns true if 'cell' has been processed. That is, if the
        corresponding row has a marker in certain column.
        """
        return self.sheet.cell(cell.row, self.metadata_columns[CONFIG_MARKER_HEADING]).value == CONFIG_MARKER

    def save(self):
        """
        Saves the current workbook.
        """
        logging.debug('Guardando workbook.')
        self.workbook.save(self.filename)
        logging.debug('Workbook guardado en «%s».', self.filename)


class LibraryCatalogueHTMLParser(HTMLParser):
    """
    Parser for web pages of items belonging to library catalogues.
    Originally created for the Biblioteca Histórica Municipal of Madrid.
    """
    def __init__(self, k_class, v_class, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.within_k_div = False
        self.within_v_div = False
        self.k_class = k_class
        self.v_class = v_class
        self.current_key = ''
        self.current_value = ''
        self.document_metadata = {}

    def handle_starttag(self, tag, attrs):
        if tag != 'div':  # Optimize by ignoring non-div tags
            return
        for attr in attrs:
            if attr[0] == 'class' and self.k_class in attr[1]:
                logging.debug('Se encontró un «%s».', self.k_class)
                self.within_k_div = True
                self.current_key = ''
            if attr[0] == 'class' and self.v_class in attr[1]:
                logging.debug('Se encontró un «%s».', self.v_class)
                self.within_v_div = True
                self.current_value = ''

    def handle_endtag(self, tag):
        if tag != 'div':  # Optimize by ignoring non-div tags
            return
        if self.within_k_div:
            self.within_k_div = False
        if self.within_v_div:
            self.within_v_div = False
            # Metadata is only stored after getting the 'titn', which is the
            # value of the metadata, so to speak.
            if not self.current_key:
                logging.debug('El «%s» está vacío.', self.k_class)
                self.current_key = '[vacío]'
            self.document_metadata[self.current_key] = self.current_value
            self.current_key = ''
            self.current_value = ''

    def handle_data(self, data):
        if self.within_k_div or self.within_v_div:
            # Clean up the received data by removing superfluous whitespace
            # characters, including newlines, carriage returns, etc.
            data = ' '.join(data.split())
            if not data:  # Ignore empty data
                return
        if self.within_k_div:
            logging.debug('Se encontró el dato «%s» «%s».', self.k_class, data)
            self.current_key += data.rstrip(':')
        if self.within_v_div:
            logging.debug('Se encontró el dato «%s» «%s».', self.v_class, data)
            if self.current_value:
                self.current_value += ' / '
            self.current_value += data

    def parse(self, contents):
        """
        Gets some library item metadata from the 'contents' HTML.
        """
        self.document_metadata.clear()
        self.feed(contents)
        self.close()
        return self.document_metadata

    def error(self, _):
        """
        This method is abstract in ParserBase, so it's better to override it.
        """
        pass


# The exception is OSError, the error code is WSAETIMEDOUT / Win10060
def retrieve_uri_contents(uri):
    """
    Resolves meta-refresh redirection for 'uri', then gets the contents and
    decodes them using the detected charset, or utf-8 if none specified.

    NOTE about charset: if no charset is specified, utf-8 is used as a
    fallback. It's better to use utf-8 rather than ascii as fallback since
    utf-8 will work too even if the page's character encoding is ascii (both
    encodings are equivalent for those codepoints), and it's very probable that
    the encoding is utf-8 if the charset is not specified in modern web pages.

    Charset is retrieved using 'get_content_charset()', which uses the HTTP
    header 'Content-Type', but there are other alternatives to get the charset
    by hand, using <meta> tags, like this:

    match = re.search(rb'<meta http-equiv="content-type" content="text/html; charset=([^"]+)"', data, re.I)
    if match:
        return match.group(1).decode('ascii')
    match = re.search(rb'<meta charset="([^"]+)"', contents, re.I)
    if match:
        return match.group(1).decode('ascii')
    return 'utf-8'
    """
    request = urlopen(uri)
    charset = request.info().get_content_charset('utf-8')
    contents = request.read().decode(charset)

    match = re.search(r'<meta http-equiv="refresh" content="[^;]+;\s*url=([^"]+)"', contents, re.I)
    if not match:
        logging.warning('No hay redirección disponible para «%s».', uri)
        return ''
    uri = urlparse(uri)
    redirected_uri = urlunparse((uri.scheme, uri.netloc, match.group(1), '', '', ''))
    logging.debug('Redirección -> «%s».', redirected_uri)

    # Get the contents of the redirected page and decode it using the proper charset.
    request = urlopen(redirected_uri)
    charset = request.info().get_content_charset('utf-8')
    logging.debug('Contenidos codificados con charset «%s».', charset)
    contents = request.read().decode(charset)
    return contents


def setup_logging(base_filename):
    """
    Sets up the program's logging system.
    """
    base_filename = os.path.expandvars(os.path.join(PROGRAM_PATH, os.path.basename(base_filename + '_log')))
    userlogfilename = base_filename + '.txt'
    debuglogfilename = base_filename + '.debug.txt'

    fmt = '%(levelname).1s: %(message)s'
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(fmt=fmt))

    fmt = '%(asctime)s |%(levelname)8s| %(message)s'
    datefmt = '%Y-%m-%d %H:%M:%S'
    userloghandler = logging.FileHandler(userlogfilename, 'w', encoding='utf-8')
    userloghandler.setLevel(logging.INFO)
    userloghandler.setFormatter(logging.Formatter(fmt=fmt, datefmt=datefmt))

    fmt = '%(asctime)s.%(msecs)04.0f |%(levelname)8s| %(message)s'
    datefmt = '%Y%m%d_%H%M%S'
    debugloghandler = logging.FileHandler(debuglogfilename, 'w', encoding='utf-8')
    debugloghandler.setLevel(logging.DEBUG)
    debugloghandler.setFormatter(logging.Formatter(fmt=fmt, datefmt=datefmt))

    logging.getLogger().setLevel(logging.NOTSET)
    logging.getLogger().addHandler(console)
    logging.getLogger().addHandler(userloghandler)
    logging.getLogger().addHandler(debugloghandler)
    logging.debug('El registro de eventos se guardará en «%s».', userlogfilename)
    logging.debug('El registro de depuración se guardará en «%s».', debuglogfilename)
    return userlogfilename


def parse_argv():
    """
    Processes the input arguments.
    """
    # The input filename should be provided automatically if the program is
    # used as a drag'n'drop target, which is the intended method of operation.
    # But the program can be also run by hand from a command prompt, so give
    # the end user a warning (well, error...) if the input filename is missing.
    if len(sys.argv) != 2:
        error('No se ha especificado un fichero de entrada para ser procesado.\n'
              'Arrastre y suelte un fichero Excel sobre el icono del programa.')
        return None

    # Well, there is at least ONE argument to the program, let's see if it's
    # the magical '--version' or some other alias.
    # In that case, output program's version.
    if sys.argv[1].lower() in ('--version', '-v', '/v'):
        import inspect
        version = None
        for line in inspect.getsourcelines(sys.modules[__name__])[0]:
            if line.startswith('# Revision'):
                version = "".join([x for x in line.split().pop() if x not in '-:+'])
                break
        if not version:
            error('Hay un error interno y no se encuentra el número de versión.')
            return None
        config = [var for var in globals() if var.startswith('CONFIG_')]
        print(f'{PROGRAM_NAME} v{"".join([x for x in version.split().pop() if x not in "-:+"])}')
        print('\nConfiguración del programa:')
        varnamelen = max(len(var) for var in config)
        for var in config:
            print(f'\t{var:<{varnamelen}} => {globals()[var]}')
        return None
    return sys.argv[1]


def process_cells(excel):
    """
    Finds cells containing URIs in the input workbook 'excel', retrieves them
    and parses the resulting HTML, getting the metadata of interest and adding
    it back to the input workbook.
    """
    error_message = None
    parser = LibraryCatalogueHTMLParser(CONFIG_K_DIV_CLASS, CONFIG_V_DIV_CLASS)
    for cell in excel.find(CONFIG_URI_PATTERN):
        if excel.is_processed(cell):
            logging.debug('El documento «%s» ya había sido procesado.', cell.value)
            continue
        logging.info('Procesando documento «%s».', cell.value)
        try:
            excel.add_metadata_for_cell(cell, parser.parse(retrieve_uri_contents(cell.value)))
        except ConnectionError as exc:
            error_message = f'Error de conexión para {cell.value}.'
        except URLError as exc:
            error_message = f'Error accediendo a «{cell.value}»: {exc.reason}.'
        finally:
            if error_message:
                logging.error(error_message)
    return error_message


def main():
    """."""
    input_filename = parse_argv()
    if not input_filename:
        return 1
    working_copy_filename = '{}_out{}'.format(*os.path.splitext(input_filename))

    #  Setup logging.
    userlogfilename = setup_logging(input_filename)
    logging.info('Registro de eventos iniciado.')
    logging.info('El fichero de entrada es «%s».', input_filename)
    logging.info('El fichero de trabajo es «%s».', working_copy_filename)

    error_message = None
    excel = None
    try:
        # A copy of the input file is created to avoid modifying the original.
        logging.debug('Copiando fichero de entrada en fichero de trabajo.')
        copy2(input_filename, working_copy_filename)
        logging.debug('Cargando fichero de trabajo.')
        excel = LibraryCatalogueExcel(working_copy_filename)
        logging.debug('Procesando celdas.')
        error_message = process_cells(excel)
    except FileNotFoundError as exc:
        error_message = 'No se encontró el fichero de entrada.'
    except PermissionError as exc:
        error_message = 'No hay permisos suficientes para '
        error_message += 'leer ' if exc.filename == input_filename else 'crear '
        error_message += 'el fichero de '
        error_message += 'entrada.' if exc.filename == input_filename else 'trabajo.'
    except (InvalidFileException, SheetTitleException):
        error_message = 'El fichero de entrada es inválido.'
    except OSError as exc:
        error_message = f'Error inesperado del sistema operativo: {exc.strerror} '
        error_message += f'[{errno.errorcode[exc.errno]}]' if exc.errno is not None else ''
        error_message += f'[Win{exc.winerror}] ' if exc.winerror is not None else ''
    finally:
        if excel:
            excel.save()
        if error_message:
            logging.error(error_message)
            error(error_message)

    logging.info('Registro de eventos finalizado.')
    logging.shutdown()
    os.startfile(userlogfilename)
    return 0 if not error_message else 1


if __name__ == '__main__':
    sys.exit(main())
